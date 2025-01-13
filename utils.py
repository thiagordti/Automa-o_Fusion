from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from tkinter.filedialog import askopenfilename
from datetime import datetime
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import tkinter as tk
import calendar
import time
import locale
import pandas as pd
import shutil
import os

locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')

def acessar_iframe(nav, tempo_espera,automacao_fusion_instance):
    """
    Acessa e muda o contexto para um iframe em uma página da web.

    Args:
        nav (WebDriver): O navegador (WebDriver) usado para interagir com a página.
        tempo_espera (int, float): Tempo, em segundos, para aguardar antes de procurar pelo iframe.

    Functionality:
        - Aguarda o tempo especificado (tempo_espera) antes de tentar acessar o iframe.
        - Espera até 180 segundos para que o iframe esteja presente no DOM da página.
        - Muda o contexto do WebDriver para o iframe localizado.

    Returns:
        None: A função realiza a troca de contexto para o iframe, sem retornar um valor.
    """
    while True:
        try:
            time.sleep(tempo_espera)  # Espera antes de mudar para o Iframe
            iframe = WebDriverWait(nav, 180).until(EC.presence_of_element_located((By.TAG_NAME, 'iframe')))  # Espera 180 segundos até o iframe aparecer
            nav.switch_to.frame(iframe)  # Troca para o iframe
            break  # Sai do loop se o comando for bem-sucedido
        except Exception:
            if not automacao_fusion_instance.handle_custom_messagebox_response():
                break

def acessar_iframe_default(nav, tempo_espera,automacao_fusion_instance, timeout=10):
    """
    Retorna ao conteúdo principal da página (fora de qualquer iframe) e acessa novamente um iframe.

    Args:
        nav (WebDriver): O navegador (WebDriver) usado para interagir com a página.
        tempo_espera (int, float): Tempo, em segundos, para aguardar antes de retornar ao conteúdo padrão.
        timeout (int, optional): Tempo máximo, em segundos, para aguardar o iframe aparecer. O padrão é 10 segundos.

    Functionality:
        - Aguarda o tempo especificado (tempo_espera) antes de trocar o contexto para o conteúdo padrão da página.
        - Muda o contexto do WebDriver para o conteúdo principal (fora de qualquer iframe).
        - Espera até o `timeout` para que o iframe esteja presente no DOM.
        - Muda o contexto do WebDriver para o iframe localizado.

    Returns:
        None: A função realiza a troca de contexto para o iframe, sem retornar um valor.
    """
    while True:
        try:
            time.sleep(tempo_espera)  # Espera antes de mudar para o conteúdo padrão
            nav.switch_to.default_content()  # Volta ao conteúdo principal da página
            iframe = WebDriverWait(nav, timeout).until(EC.presence_of_element_located((By.TAG_NAME, 'iframe')))  # Espera até que o iframe esteja presente
            nav.switch_to.frame(iframe)  # Troca para o iframe
            break  # Sai do loop se o comando for bem-sucedido
        except Exception:
            if not automacao_fusion_instance.handle_custom_messagebox_response():
                break

def clicar_elemento(nav, elemento, tipo,automacao_fusion_instance):
    """
    Localiza e clica em um elemento na página web utilizando JavaScript.

    Args:
        nav (WebDriver): O navegador (WebDriver) usado para interagir com a página.
        elemento (str): O seletor do elemento a ser localizado na página.
        tipo (By): O tipo de seletor (e.g., By.ID, By.CLASS_NAME, etc.) usado para localizar o elemento.

    Functionality:
        - Tenta localizar o elemento especificado utilizando `WebDriverWait`, aguardando até 30 segundos para ele aparecer.
        - Clica no elemento utilizando um comando JavaScript para garantir a execução do clique.
        - Se o elemento não for encontrado após o tempo de espera, exibe um alerta através de uma janela Tkinter, informando o usuário para interagir manualmente.

    Returns:
        None: A função tenta localizar e clicar no elemento, e em caso de falha, exibe uma mensagem de alerta ao usuário.

    Raises:
        Exibe um alerta ao usuário se o elemento não for encontrado dentro do tempo de espera.
    """
    while True:
        try:
            obj = WebDriverWait(nav, 10).until(EC.presence_of_element_located((tipo, elemento)))  # Aguarda 10 segundos até o elemento carregar
            nav.execute_script("arguments[0].click();", obj)  # Clica no objeto utilizando JavaScript
            break  # Sai do loop se o comando for bem-sucedido
        except Exception:
            if not automacao_fusion_instance.handle_custom_messagebox_response():
                break
            
def clicar_elemento_dinamico(nav,automacao_fusion_instance):
    """
    Localiza e clica em um elemento cujo ID é dinâmico, utilizando XPath para identificar pelo padrão do ID.

    Args:
        nav (WebDriver): O navegador (WebDriver) usado para interagir com a página.

    Functionality:
        - Tenta localizar o elemento especificado utilizando `WebDriverWait`, aguardando até 30 segundos para ele estar visível e clicável.
        - Clica no elemento utilizando um comando JavaScript para garantir a execução do clique.
        - Se o elemento não for encontrado após o tempo de espera, exibe um alerta através de uma janela Tkinter, informando o usuário para interagir manualmente.

    Returns:
        None: A função tenta localizar e clicar no elemento, e em caso de falha, exibe uma mensagem de alerta ao usuário.
    """
    while True:
        try:
            # XPath que localiza um elemento cujo ID começa com 'ui-id-' e que é um link (a tag)
            xpath = "//a[starts-with(@id, 'ui-id-')]"
            
            # Aguarda até que o elemento esteja visível e clicável (até 30 segundos)
            obj = WebDriverWait(nav, 30).until(EC.element_to_be_clickable((By.XPATH, xpath)))
            
            nav.execute_script("arguments[0].scrollIntoView(true);", obj)  # Garante que o elemento esteja visível na tela
            nav.execute_script("arguments[0].click();", obj)  # Clica no objeto utilizando JavaScript
            break  # Sai do loop se o clique for bem-sucedido
        except Exception:
            if not automacao_fusion_instance.handle_custom_messagebox_response():
                break

def clicar_elemento_rustico(nav, elemento, tipo,automacao_fusion_instance):
    """
    Localiza e clica em um elemento na página web usando o método tradicional do Selenium.

    Args:
        nav (WebDriver): O navegador (WebDriver) usado para interagir com a página.
        elemento (str): O seletor do elemento a ser localizado na página.
        tipo (By): O tipo de seletor (e.g., By.ID, By.CLASS_NAME, etc.) usado para localizar o elemento.

    Functionality:
        - Aguarda até 60 segundos para que o elemento especificado esteja presente no DOM.
        - Utiliza o método padrão do Selenium `find_element` para localizar e clicar no elemento.
        - Se o elemento não for encontrado após o tempo de espera, exibe um alerta através de uma janela Tkinter, informando o usuário para interagir manualmente.

    Returns:
        None: A função tenta localizar e clicar no elemento, e em caso de falha, exibe uma mensagem de alerta ao usuário.

    Raises:
        Exibe um alerta ao usuário se o elemento não for encontrado dentro do tempo de espera.
    """
    while True:
        try:
            WebDriverWait(nav, 15).until(EC.presence_of_element_located((tipo, elemento)))  # Aguarda 15 segundos até o elemento carregar
            nav.find_element(tipo, elemento).click()  # Clica no elemento usando o método padrão do Selenium
            break  # Sai do loop se o comando for bem-sucedido
        except Exception:
            if not automacao_fusion_instance.handle_custom_messagebox_response():
                break

def enviarkey_elemento(nav, elemento, tipo, texto,automacao_fusion_instance):
    """
    Localiza um elemento na página web e envia um texto para ele.

    Args:
        nav (WebDriver): O navegador (WebDriver) usado para interagir com a página.
        elemento (str): O seletor do elemento a ser localizado na página.
        tipo (By): O tipo de seletor (e.g., By.ID, By.CLASS_NAME, etc.) usado para localizar o elemento.
        texto (str): O texto que será enviado para o elemento localizado.

    Functionality:
        - Aguarda até 60 segundos para que o elemento especificado esteja presente no DOM.
        - Utiliza o método padrão do Selenium `find_element` para localizar o elemento e enviar o texto fornecido.
        - Se o elemento não for encontrado após o tempo de espera, exibe um alerta através de uma janela Tkinter, informando o usuário para interagir manualmente.

    Returns:
        None: A função envia texto para o elemento localizado, e em caso de falha, exibe uma mensagem de alerta ao usuário.

    Raises:
        Exibe um alerta ao usuário se o elemento não for encontrado dentro do tempo de espera.
    """
    while True:
        try:
            WebDriverWait(nav, 60).until(EC.presence_of_element_located((tipo, elemento)))  # Aguarda 60 segundos até o elemento carregar
            nav.find_element(tipo, elemento).send_keys(texto)  # Envia o texto para o elemento
            break  # Sai do loop se o comando for bem-sucedido
        except Exception:
            if not automacao_fusion_instance.handle_custom_messagebox_response():
                break

def primeiro_e_ultimo_dia_do_mes(ano, mes):
    """
    Retorna o primeiro e o último dia de um mês específico no formato 'dd/mm/aaaa'.

    Args:
        ano (int): O ano desejado.
        mes (int): O número do mês desejado (1 a 12).

    Functionality:
        - Calcula o primeiro dia do mês especificado.
        - Calcula o último dia do mês com base no número de dias do mês fornecido.
        - Formata ambos os dias no formato 'dd/mm/aaaa'.

    Returns:
        tuple: Uma tupla contendo o primeiro e o último dia do mês no formato 'dd/mm/aaaa'.
    """
    primeiro_dia = datetime(ano, mes, 1)  # Primeiro dia do mês
    ultimo_dia = datetime(ano, mes, calendar.monthrange(ano, mes)[1])  # Último dia do mês
    # Formatar as datas no formato dd/mm/aaaa
    primeiro_dia_formatado = primeiro_dia.strftime('%d/%m/%Y')
    ultimo_dia_formatado = ultimo_dia.strftime('%d/%m/%Y')
    return primeiro_dia_formatado, ultimo_dia_formatado

def copiar_linha_ativa(df, destino, sheet_name, linha, texto_adicional=None):
    """
    Copia uma linha ativa de um DataFrame para uma planilha Excel e insere dados adicionais, como a data atual e texto específico, dependendo da aba.

    Args:
        df (DataFrame): O DataFrame de onde a linha ativa será copiada.
        destino (str): O caminho do arquivo Excel de destino.
        sheet_name (str): O nome da aba (sheet) onde os dados serão copiados.
        linha (int): O índice da linha a ser copiada do DataFrame.
        texto_adicional (str, optional): Texto a ser inserido na coluna específica se o sheet_name for 'Novo'.

    Functionality:
        - Seleciona a linha especificada do DataFrame que não está vazia.
        - Abre a planilha de destino e localiza a próxima linha vazia.
        - Copia o conteúdo da linha selecionada para a planilha.
        - Se a aba for 'Novo', insere um texto adicional e a data atual em colunas específicas.
        - Caso contrário, insere apenas a data atual em uma coluna diferente.
        - Salva as alterações na planilha de destino.

    Returns:
        None: A função salva as alterações diretamente no arquivo Excel.

    Raises:
        FileNotFoundError: Se o arquivo de destino não for encontrado.
    """
    linha_ativa = df.iloc[[linha]].dropna(how='all')  # Seleciona a linha ativa específica (não vazia)
    book = load_workbook(destino)  # Tenta carregar a planilha de destino existente:
    sheet = book[sheet_name]
    next_row = sheet.max_row + 1# Encontra a próxima linha vazia na planilha de destino
    for r_idx, row in enumerate(dataframe_to_rows(linha_ativa, index=False, header=False), start=next_row):# Adiciona a linha ativa à planilha de destino
        for c_idx, value in enumerate(row, 1):
            sheet.cell(row=r_idx, column=c_idx, value=value)
    if sheet_name == 'Novo':
        cob_column_index = 34  # Index da coluna COB, alterar manualmente caso planilha seja modificada!!
        sheet.cell(row=next_row, column=cob_column_index, value=texto_adicional)# Adiciona o texto na última coluna da nova linha
        hoje = datetime.today().strftime('%d/%m/%Y')# Pega a data de hoje
        sheet.cell(row=next_row, column=cob_column_index + 1, value=hoje)# Adiciona a data de hoje na coluna seguinte
    else:
        dia_column_index = 29
        hoje = datetime.today().strftime('%d/%m/%Y')# Pega a data de hoje
        sheet.cell(row=next_row, column=dia_column_index, value=hoje)# Adiciona a data de hoje na coluna  
    
    book.save(destino)# Salva o arquivo de destino

def copiar_para_planilha(local_destino, local_origem):
    """
    Copia um arquivo de origem para um destino especificado, criando diretórios intermediários se necessário.

    Args:
        local_destino (str): O caminho do arquivo de destino, incluindo o nome do arquivo. Diretórios intermediários serão criados se não existirem.
        local_origem (str): O caminho do arquivo de origem que será copiado para o destino.

    Functionality:
        - Cria os diretórios intermediários do caminho de destino, se ainda não existirem.
        - Copia o arquivo de origem para o caminho de destino, preservando a data e hora de modificação original do arquivo.

    Returns:
        None: A função não retorna valores, apenas realiza a cópia do arquivo.

    Raises:
        FileNotFoundError: Se o arquivo de origem não for encontrado.
        PermissionError: Se houver problemas de permissão ao criar diretórios ou copiar o arquivo.
    """
    os.makedirs(os.path.dirname(local_destino), exist_ok=True)
    shutil.copy2(local_origem, local_destino)

def enviarkey_java(nav, element_name, value,automacao_fusion_instance):
    """
    Insere um valor em um campo de entrada HTML via JavaScript e dispara eventos associados, evitando possíveis interferências de máscaras de entrada.

    Args:
        nav (webdriver): O objeto WebDriver utilizado para interagir com o navegador.
        element_name (str): O nome do elemento HTML (atributo `name`) onde o valor será inserido.
        value (str): O valor que será inserido no campo.

    Functionality:
        - Aguarda até que o elemento de entrada com o nome especificado esteja presente na página.
        - Utiliza JavaScript para definir o valor diretamente no campo de entrada, evitando interferências de scripts de máscaras.
        - Dispara eventos 'input' no campo para garantir que o valor inserido seja reconhecido corretamente pelo JavaScript da página.
        - Exibe uma mensagem de alerta, caso o elemento não seja encontrado, permitindo a continuação manual.

    Raises:
        TimeoutException: Se o elemento não for encontrado dentro do tempo limite especificado (60 segundos).
    
    Returns:
        None: A função não retorna valores, apenas interage com o navegador.
    """
    while True:
        try:
            WebDriverWait(nav, 60).until(EC.presence_of_element_located((By.NAME, element_name)))
            script = f"document.getElementsByName('{element_name}')[0].value='{value}';" # Simula a entrada de dados via JavaScript para evitar interferência da máscara de entrada
            nav.execute_script(script)
            script = f"""
            var input = document.getElementsByName('{element_name}')[0];
            var event = new Event('input', {{ bubbles: true }});
            input.dispatchEvent(event);
            """
            nav.execute_script(script) # Dispara eventos para que o script de máscara possa processar o novo valor
            break  # Sai do loop se o clique for bem-sucedido
        except Exception:
            if not automacao_fusion_instance.handle_custom_messagebox_response():
                break      

def selecionar_arquivo():
    """
    Abre uma caixa de diálogo para o usuário selecionar um arquivo e retorna o caminho do arquivo selecionado.

    Returns:
        str: O caminho completo do arquivo selecionado pelo usuário.

    Functionality:
        - Utiliza uma caixa de diálogo para permitir que o usuário navegue e selecione um arquivo.
        - O título da caixa de diálogo é "Selecione a Planilha COB!".

    Raises:
        Nenhuma: A função não levanta exceções explícitas, mas pode retornar uma string vazia se o usuário cancelar a seleção.

    Notes:
        - Se o usuário cancelar a seleção, a função retornará uma string vazia.
    """
    caminho_arquivo = askopenfilename(title="Selecione a Planilha COB!") # Solciita o usuario selecionar a planilha!
    return caminho_arquivo

def texto_elemento(nav,elemento,tipo):
    """
    Extrai o texto de um elemento HTML localizado na página e retorna a parte do texto antes do caractere hífen ('-').

    Args:
        nav (webdriver): O objeto WebDriver utilizado para interagir com o navegador.
        elemento (str): O identificador do elemento HTML a ser localizado (pode ser um nome, ID, classe, etc.).
        tipo (By): O método de localização do elemento (por exemplo, By.NAME, By.ID, By.CLASS_NAME).

    Returns:
        str: O texto extraído do elemento, antes do primeiro caractere hífen ('-').

    Functionality:
        - Aguarda até que o elemento especificado esteja presente na página.
        - Obtém o texto do elemento localizado.
        - Divide o texto no caractere hífen ('-') e retorna a primeira parte.

    Raises:
        TimeoutException: Se o elemento não for encontrado dentro do tempo limite especificado (60 segundos).

    Notes:
        - Se o texto do elemento não contiver um hífen, a função retornará o texto completo.
    """
    obj = WebDriverWait(nav, 60).until(EC.presence_of_element_located((tipo, elemento))) # Aguarda 60 segundos até o elemento carregar
    texto = obj.text
    return texto.split('-')[0]

def esperar_elementos_carregar(nav, timeout=60):
    """
    Aguarda até que pelo menos um dos elementos especificados esteja presente na página ou até que uma mensagem de "nenhum resultado" seja exibida.

    Args:
        nav (webdriver): O objeto WebDriver utilizado para interagir com o navegador.
        timeout (int, opcional): O tempo máximo de espera em segundos. O padrão é 60 segundos.

    Returns:
        None: A função não retorna valores, apenas aguarda a condição especificada ser atendida.

    Functionality:
        - Espera até que pelo menos um elemento com a classe "item" esteja presente na página.
        - Ou, espera até que um elemento contendo a mensagem "Sua Caixa de Entrada está vazia" seja exibido.
        - O tempo máximo de espera pode ser especificado pelo parâmetro `timeout`.

    Raises:
        TimeoutException: Se a condição especificada não for atendida dentro do tempo limite definido.

    Notes:
        - A função utiliza uma expressão lambda para verificar a presença dos elementos ou a mensagem de "nenhum resultado".
    """
    WebDriverWait(nav, timeout).until(
        lambda driver: len(driver.find_elements(By.CLASS_NAME, "item")) > 0 or 
                        len(driver.find_elements(By.XPATH, '//div[contains(@class, "no-results-default-boxes") and contains(@class, "ng-scope") and contains(., "Sua Caixa de Entrada está vazia")]')) > 0
    )

def enviar_emails(nav, linha, click, campo, planilha, tempo_espera, automacao_fusion_instance):
    """
    Envia e-mails para endereços listados em uma célula de uma planilha, interagindo com elementos da interface web via WebDriver.

    Args:
        nav (webdriver): O objeto WebDriver utilizado para interagir com o navegador.
        linha (int): O índice da linha na planilha onde os endereços de e-mail estão localizados.
        click (str): O identificador (XPath) do elemento a ser clicado para iniciar o envio de e-mails.
        campo (str): O identificador (ID) do campo de entrada onde os e-mails serão inseridos.
        planilha (pandas.DataFrame): O DataFrame que contém os e-mails, com uma coluna chamada 'EMAILS' que lista os e-mails separados por '/'.
        tempo_espera (float): O tempo de espera para acessar iframes e elementos.
        automacao_fusion_instance (AutomacaoFusion): A instância da classe AutomacaoFusion.

    Returns:
        None: A função não retorna valores, apenas executa ações na interface web.

    Functionality:
        - Obtém a lista de e-mails da planilha, separando-os pelo caractere '/'.
        - Clica no elemento especificado para iniciar o processo de envio de e-mails.
        - Itera sobre a lista de e-mails, acessa o iframe necessário, insere o e-mail no campo apropriado e realiza ações adicionais para enviar o e-mail.
        - Clica no botão "OK" após cada envio e, finalmente, fecha as janelas com o botão "Cancelar".

    Raises:
        Exception: Se ocorrer um erro ao clicar em elementos ou enviar dados, a função pode interromper a execução e exibir um alerta.

    Notes:
        - A função utiliza as funções auxiliares `clicar_elemento`, `acessar_iframe_default`, `enviarkey_elemento` e `clicar_elemento_rustico` para interagir com a interface da web.
        - A função assume que todos os e-mails estão separados por '/' e que os elementos de interface estão corretamente identificados pelos parâmetros fornecidos.
    """
    email = planilha.iloc[linha]['EMAILS']  # recebe e-mails da planilha (Os mesmos devem ser separados por uma '/')
    lst_email = email.split('/')  # Transforma os e-mails recebidos em lista, separador '/'
    clicar_elemento(nav, click, By.CSS_SELECTOR, automacao_fusion_instance)  # Itens novos e-mails
    for i in range(len(lst_email)):
        acessar_iframe_default(nav, tempo_espera, automacao_fusion_instance)  # Acessa Iframe dos e-mails
        enviarkey_elemento(nav, campo, By.ID, lst_email[i], automacao_fusion_instance)  # Envia e-mail
        clicar_elemento_rustico(nav, 'form_container', By.ID, automacao_fusion_instance)  # Clica no container para o e-mail carregar
        clicar_elemento(nav, '//*[@id="dibButtons"]/input[1]', By.XPATH, automacao_fusion_instance)  # Botão Ok
    acessar_iframe_default(nav, tempo_espera, automacao_fusion_instance)  # Acessa Iframe dos e-mails
    clicar_elemento(nav, 'cancelButtonModal', By.ID, automacao_fusion_instance)  # Botão Cancelar, para fechar janelas!

def tratar_cnpj(cnpj):
    """
    Formata e limpa um CNPJ, removendo pontuações e zeros à esquerda.

    Args:
        cnpj (str ou int): O CNPJ a ser tratado, que pode estar no formato com pontuações (ex.: '12.345.678/0001-95') ou sem pontuações.

    Returns:
        str: O CNPJ formatado e limpo, sem pontuações e zeros à esquerda.

    Functionality:
        - Converte o CNPJ para uma string, se necessário.
        - Remove quaisquer pontuações presentes no CNPJ, como '.', '-', e '/'.
        - Converte o CNPJ limpo para um número inteiro para remover zeros à esquerda e, em seguida, converte de volta para string.

    Notes:
        - A função assume que o CNPJ fornecido é válido e pode conter pontuações comuns.
        - O CNPJ formatado é retornado como uma string sem pontuações e sem zeros à esquerda.
    """
    cnpj = str(cnpj)  # Converte o CNPJ para string, caso não seja
    if '.' in cnpj or '-' in cnpj or '/' in cnpj:# Verifica se o CNPJ contém '.', '-', '/'
        cnpj_limpo = cnpj.replace('.', '').replace('/', '').replace('-', '')# Remove pontuações se existirem
    else:
        cnpj_limpo = cnpj  # Se não tiver pontuações, já está limpo
    cnpj_formatado = str(int(cnpj_limpo))# Converte para inteiro para remover zeros à esquerda, depois para string novamente
    return cnpj_formatado

def enviar_anexo(nav, linha, click, element, status, descr, planilha, caminho, tempo_espera, automacao_fusion_instance):
    """
    Envia anexos e suas descrições para uma interface web, iterando sobre os anexos listados em uma planilha.

    Args:
        nav (webdriver): O objeto WebDriver utilizado para interagir com o navegador.
        linha (int): O índice da linha na planilha que contém as informações dos anexos.
        click (str): O identificador (XPath) do elemento a ser clicado para iniciar o envio de anexos.
        element (str): O identificador (ID) do campo onde o caminho do arquivo deve ser inserido.
        status (str): O identificador (XPath) para verificar se o anexo foi carregado com sucesso.
        descr (str): O identificador (ID) do campo onde a descrição do anexo deve ser inserida.
        planilha (pandas.DataFrame): O DataFrame que contém as informações dos anexos e descrições.
        caminho (str): O caminho base onde os arquivos a serem anexados estão localizados.
        tempo_espera (float): O tempo de espera para carregar os elementos.
        automacao_fusion_instance (AutomacaoFusion): A instância da classe AutomacaoFusion.

    Returns:
        None: A função não retorna valores, apenas executa ações na interface web.

    Functionality:
        - Itera sobre dois possíveis anexos (ou mais, se alterado).
        - Verifica se o anexo está presente na planilha; se estiver, clica no elemento para iniciar o envio.
        - Acessa o iframe necessário para o envio do anexo.
        - Insere o caminho do arquivo no campo apropriado e aguarda o carregamento do item.
        - Insere a descrição do anexo e confirma o envio clicando no botão correspondente.
        - Volta para o conteúdo padrão após cada envio.

    Notes:
        - A função assume que os anexos e descrições estão nas colunas 'ARQUIVO1', 'ARQUIVO2', 'DESCRICAO1', e 'DESCRICAO2' da planilha.
        - O caminho do arquivo é formatado a partir do caminho base fornecido, substituindo a parte final do caminho.
        - A função faz uso das funções auxiliares `clicar_elemento`, `acessar_iframe`, e `enviarkey_elemento` para interagir com a interface da web.
    """
    for anexo in range(2): 
        if pd.isna(planilha.iloc[linha][f'ARQUIVO{anexo+1}']):
            pass
        else:
            clicar_elemento(nav, click, By.XPATH, automacao_fusion_instance)  # Clica no anexo para enviar arquivo
            acessar_iframe(nav, tempo_espera, automacao_fusion_instance)  # Acesso Iframe
            enviarkey_elemento(nav, element, By.ID, fr"{caminho[:-16]}Arquivos\{planilha.iloc[linha][f'ARQUIVO{anexo+1}']}", automacao_fusion_instance)  # Envia o anexo
            while len(nav.find_elements(By.XPATH, status)) == 0:  # Loop para aguardar a lista de itens carregar, se a lista não carregar a pesquisa não funciona!
                time.sleep(1)
            time.sleep(1)
            enviarkey_elemento(nav, descr, By.ID, planilha.iloc[linha][f'DESCRICAO{anexo+1}'], automacao_fusion_instance)  # Envia a descrição
            clicar_elemento(nav, '//*[@id="dibButtons"]/input[1]', By.XPATH, automacao_fusion_instance) 
            nav.switch_to.default_content()  # Volta para o inicio

def opcoes_pagamento(nav,selec,seta):
    """
    Seleciona opções de pagamento em uma interface web.

    Args:
        nav (webdriver): O objeto WebDriver utilizado para interagir com o navegador.
        selec (str): O identificador (XPath) do elemento que deve ser clicado para abrir as opções de pagamento.
        seta (str): O identificador (ID) do elemento que deve ser clicado para selecionar uma opção específica de pagamento.

    Returns:
        None: A função não retorna valores, apenas executa ações na interface web.

    Functionality:
        - Executa um loop para selecionar duas opções de pagamento.
        - Clica no elemento especificado por `selec` para abrir as opções de pagamento.
        - Clica no elemento especificado por `seta` para selecionar uma das opções disponíveis.

    Notes:
        - O loop é executado duas vezes, o que indica que a função está projetada para selecionar duas opções de pagamento consecutivamente.
        - A função assume que os elementos identificados por `selec` e `seta` estão presentes na interface da web e são funcionais.
    """
    for i in range(2):#Loop para selecionar as opções de pagamento
        nav.find_element(By.XPATH, selec).click()
        nav.find_element(By.ID, seta).click()  

def clicar_porcentagem(nav, contador, linha, planilha, tempo_espera, automacao_fusion_instance):
    """
    Clica em itens com base em um contador e envia um número de contrato para uma interface web, salvando as alterações.

    Args:
        nav (webdriver): O objeto WebDriver utilizado para interagir com o navegador.
        contador (int): O número de itens a serem clicados, determinado pela soma do contador.
        linha (int): O índice da linha na planilha que contém o número do contrato.
        planilha (pandas.DataFrame): O DataFrame que contém os dados dos contratos.
        tempo_espera (float): O tempo de espera para carregar os elementos.
        automacao_fusion_instance (AutomacaoFusion): A instância da classe AutomacaoFusion.

    Returns:
        None: A função não retorna valores, apenas executa ações na interface web.

    Functionality:
        - Executa um loop baseado no valor de `contador` para clicar em itens específicos na interface.
        - Acessa o iframe primário e salva as alterações após clicar em cada item.
        - Envia o número do contrato obtido da planilha para o campo apropriado na interface.
        - Clica no número do contrato para selecioná-lo e, em seguida, salva as alterações.
        - Volta para o conteúdo padrão após salvar todas as alterações.

    Notes:
        - A função assume que os elementos identificados pelos seletores XPath e ID estão presentes e funcionais na interface da web.
        - O índice `i` no XPath dos itens é baseado em um sistema que conta a partir de 0 (ou seja, o primeiro item tem índice 0).
        - A função utiliza as funções auxiliares `clicar_elemento_rustico`, `acessar_iframe_default`, e `enviarkey_elemento` para interagir com a interface da web.
    """
    for i in range(contador):  # Baseado na soma do Contador clica nos itens
        clicar_elemento_rustico(nav, f'//*[@id="{i}"]/td[2]', By.XPATH, automacao_fusion_instance)  # Clica no Item baseado nos indices (No fusion o indice 0 conta!)!!
        acessar_iframe_default(nav, tempo_espera, automacao_fusion_instance)  # Acessa Iframe primario
        clicar_elemento(nav, 'action.save', By.NAME, automacao_fusion_instance)  # Clica para salvar.
        acessar_iframe_default(nav, tempo_espera, automacao_fusion_instance)  # Acessa Iframe primario
    clicar_elemento(nav, 'action.save', By.NAME, automacao_fusion_instance)  # Clica para salvar.
    nav.switch_to.default_content()  # Volta para o inicio

def dados_rateio(nav, linha, cod_filial, cod_uo, planilha, tempo_espera, automacao_fusion_instance):
    """
    Preenche dados de rateio em um formulário de interface web com base nas informações fornecidas.

    Args:
        nav (webdriver): O objeto WebDriver utilizado para interagir com o navegador.
        linha (int): O índice da linha na planilha que contém dados de rateio.
        cod_filial (str): Código da filial a ser enviado no formulário.
        cod_uo (str): Código da Unidade Organizacional (UO) a ser enviado no formulário.
        planilha (pandas.DataFrame): O DataFrame que contém os dados de rateio, incluindo a classe de valor.
        tempo_espera (float): O tempo de espera para carregar os elementos.
        automacao_fusion_instance (AutomacaoFusion): A instância da classe AutomacaoFusion.

    Returns:
        None: A função não retorna valores, apenas executa ações na interface web.

    Functionality:
        - Clica em um item no menu para abrir o campo de produtos.
        - Acessa o iframe da pesquisa de produtos.
        - Seleciona uma opção de pagamento do campo apropriado.
        - Clica para abrir o campo de pesquisa.
        - Acessa o iframe da pesquisa.
        - Clica para abrir o filtro.
        - Acessa o iframe do filtro.
        - Envia a Classe de Valor, o Código da Filial e o Código da UO para os campos apropriados.

    Notes:
        - A função assume que os elementos identificados pelos seletores XPath e ID estão presentes e funcionais na interface da web.
        - O `planilha` deve conter uma coluna chamada 'CLASSE DE VALOR' com valores numéricos.
        - A função utiliza as funções auxiliares `clicar_elemento`, `acessar_iframe_default`, `opcoes_pagamento`, e `enviarkey_elemento` para interagir com a interface da web.
    """
    clicar_elemento(nav, '//*[@id="menu_bar_FINFFCobFaturamentoVariavelCentroDeResultadosXFilialXValor"]/li[1]', By.XPATH, automacao_fusion_instance)  # Clica para abrir campo de produtos
    acessar_iframe_default(nav, tempo_espera, automacao_fusion_instance)  # Acessa Iframe da Pesquisa de produtos
    opcoes_pagamento(nav, '//*[@id="mul_dadosDaCobranca__dadosDoFaturamentoVariavel__dadosDoRateio__formaDeEntradaDosRecursos_ori"]/option[1]', 'move_this_right_mul_dadosDaCobranca__dadosDoFaturamentoVariavel__dadosDoRateio__formaDeEntradaDosRecursos')  # Loop para selecionar as opções de pagamento
    clicar_elemento(nav, 'id_dadosDaCobranca__dadosDoFaturamentoVariavel__dadosDoRateio__UOCRProtheus___anchor', By.ID, automacao_fusion_instance)  # Clica para abrir campo de pesquisa
    acessar_iframe_default(nav, tempo_espera, automacao_fusion_instance)  # Acessa Iframe da Pesquisa
    clicar_elemento(nav, '//*[@id="menu_bar_EXTERNOProtheusAmarracaoContabil"]/li', By.XPATH, automacao_fusion_instance)  # Clica para abrir filtro
    acessar_iframe_default(nav, tempo_espera, automacao_fusion_instance)  # Acessa Iframe do Filtro
    enviarkey_elemento(nav, 'var_codclvlr__', By.NAME, str(int(planilha.iloc[linha]['CLASSE DE VALOR'])), automacao_fusion_instance)  # Envia Classe de valor Cliente
    enviarkey_elemento(nav, 'var_codfilialprotheus__', By.NAME, cod_filial, automacao_fusion_instance)  # Envia COD FILIAL - PADRÃO
    enviarkey_elemento(nav, 'var_coduo__', By.NAME, cod_uo, automacao_fusion_instance)  # Envia COD UO - PADRÃO

def iniciar_navegador():
    """
    Inicializa e configura uma instância do navegador Chrome para automação.

    Returns:
        webdriver.Chrome: O objeto WebDriver para o navegador Chrome, configurado e pronto para uso.

    Functionality:
        - Configura o ChromeOptions para o navegador Chrome.
        - Adiciona uma opção experimental para manter o navegador aberto após a execução do script.
        - Inicializa uma nova instância do navegador Chrome com as opções configuradas.
        - Navega para a URL especificada 'https://fusion.fiemg.com.br/fusion/portal'.
        - Maximiza a janela do navegador para garantir que a interface esteja completamente visível.

    Notes:
        - A função pressupõe que o driver do Chrome (`chromedriver`) está corretamente instalado e disponível no PATH do sistema.
        - A opção `detach` permite que o navegador continue aberto após a conclusão do script, útil para depuração e verificação manual.
    """
    options = webdriver.ChromeOptions()
    options.add_experimental_option("detach", True)# Para o mesmo não fechar apos execução
    navegador = webdriver.Chrome(options=options)# Executa o navegador
    navegador.get('https://fusion.fiemg.com.br/fusion/portal')
    navegador.maximize_window()# Maximiza a janela do navegador
    return navegador
    
def esperar_alerta(nav, cob, aba_original,planilha, local_destino,nome_guia,linha,texto_adicional=None):
    """
    Aguarda a exibição de um alerta de sucesso ou falha em uma página web e executa ações baseadas no tipo de alerta.

    Args:
        nav (webdriver.Chrome): Instância do WebDriver para o navegador Chrome.
        cob (str): Identificador do COB (código de operação bancária) para mensagens de erro.
        aba_original (str): Identificador da aba original para retornar após o processamento.
        planilha (pd.DataFrame): DataFrame contendo os dados para copiar para a planilha de destino.
        local_destino (str): Caminho do arquivo de destino onde os dados serão copiados.
        nome_guia (str): Nome da guia para determinar o destino dos dados copiados ('Medição' ou 'Novo').
        linha (int): Índice da linha na planilha contendo os dados a serem copiados.
        texto_adicional (str, opcional): Texto adicional para ser incluído na guia 'Novo'. Padrão é None.

    Behavior:
        - Espera até 30 segundos para que um alerta apareça na página.
        - Verifica se o alerta é de sucesso com a classe 'alert-success'.
        - Se o alerta for de sucesso e `nome_guia` for 'Medição', chama a função `copiar_linha_ativa` para copiar os dados para a guia 'Medição'.
        - Se o alerta for de sucesso e `nome_guia` for 'Novo', chama a função `copiar_linha_ativa` para copiar os dados para a guia 'Novo' com texto adicional, se fornecido.
        - Fecha a aba do navegador e retorna à aba original após o processamento.
        - Se o alerta não for de sucesso, imprime uma mensagem de falha e também fecha a aba e retorna à aba original.

    Notes:
        - Certifique-se de que o seletor CSS `.alert` seja apropriado para capturar o alerta desejado.
        - A função pressupõe que as funções `copiar_linha_ativa` e o WebDriver estão corretamente configurados.
    """
    alert = WebDriverWait(nav, 30).until(EC.presence_of_element_located((By.CSS_SELECTOR, '.alert')))# Espera que qualquer alerta de sucesso ou perigo apareça
    if 'alert-success' in alert.get_attribute('class'):# Verifica a classe do alerta
        if nome_guia == 'Medição':
            copiar_linha_ativa(planilha, local_destino, 'Medição', linha)
        elif nome_guia == 'Novo':
            copiar_linha_ativa(planilha, local_destino, 'Novo', linha,texto_adicional)
        print('Cob executado com sucesso!!')
        nav.close()  # Fecha a aba após o alerta carregar
        nav.switch_to.window(aba_original)  # Volta para a aba original
    else:
        print(f"O {cob} apresentou falha ao enviar, não foi incluído na planilha Histórico!!")
        nav.close()  # Fecha a aba após o alerta carregar
        nav.switch_to.window(aba_original)  # Volta para a aba original

