from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException
from selenium.webdriver.support.ui import WebDriverWait
from tkinter.filedialog import askopenfilename
from datetime import datetime
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from getpass import getpass
import tkinter as tk
import calendar
import time
import locale
import pandas as pd
import shutil
import os

locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')

def acessar_iframe(nav):
    iframe = WebDriverWait(nav, 180).until(EC.presence_of_element_located((By.TAG_NAME, 'iframe'))) # Espera 180 segundos até o iframe aparecer!
    nav.switch_to.frame(iframe) #Troca para o iframe

def acessar_iframe_default(nav, timeout=10, wait_before_switch=1, max_attempts=10):
    time.sleep(wait_before_switch)# Espera antes de mudar para o conteúdo padrão
    nav.switch_to.default_content()
    iframe = WebDriverWait(nav, timeout).until(EC.presence_of_element_located((By.TAG_NAME, 'iframe')))# Espera até que o iframe esteja presente
    nav.switch_to.frame(iframe)# Troca para o iframe
    return 

def clicar_elemento(nav,elemento,tipo):
    obj = WebDriverWait(nav, 180).until(EC.presence_of_element_located((tipo, elemento))) # Aguarda 180 segundos até o elemento carregar
    navegador.execute_script("arguments[0].click();", obj) # Clica no objeto utilizando paramentros JavaScript.

def clicar_elemento_rustico(nav,elemento,tipo):
    WebDriverWait(nav, 180).until(EC.presence_of_element_located((tipo, elemento))) # Aguarda 180 segundos até o elemento carregar
    nav.find_element(tipo, elemento).click()# Clicar na Pesquisa da forma tradicional do Selenium!

def enviarkey_elemento(nav,elemento,tipo,texto):
    WebDriverWait(nav, 180).until(EC.presence_of_element_located((tipo, elemento))) # Aguarda 180 segundos até o elemento carregar
    navegador.find_element(tipo, elemento).send_keys(texto) # envia dados para o elemento

def primeiro_e_ultimo_dia_do_mes(ano, mes):
    primeiro_dia = datetime(ano, mes, 1)# Primeiro dia do mês
    ultimo_dia = datetime(ano, mes, calendar.monthrange(ano, mes)[1])# Último dia do mês
    # Formatar as datas no formato dd/mm/aaaa
    primeiro_dia_formatado = primeiro_dia.strftime('%d/%m/%Y')
    ultimo_dia_formatado = ultimo_dia.strftime('%d/%m/%Y')
    return primeiro_dia_formatado, ultimo_dia_formatado

def copiar_linha_ativa(df, destino, sheet_name, linha, texto_adicional=None):
    linha_ativa = df.iloc[[linha]].dropna(how='all')  # Seleciona a linha ativa específica (não vazia)
    book = load_workbook(destino)  # Tenta carregar a planilha de destino existente
    if sheet_name in book.sheetnames:
        sheet = book[sheet_name]
    else:
        sheet = book.create_sheet(sheet_name)
    next_row = sheet.max_row + 1# Encontra a próxima linha vazia na planilha de destino
    for r_idx, row in enumerate(dataframe_to_rows(linha_ativa, index=False, header=False), start=next_row):# Adiciona a linha ativa à planilha de destino
        for c_idx, value in enumerate(row, 1):
            sheet.cell(row=r_idx, column=c_idx, value=value)
    sheet.cell(row=next_row, column=sheet.max_column + 1, value=texto_adicional)# Adiciona o texto na última coluna da nova linha
    book.save(destino)# Salva o arquivo de destino

def copiar_para_planilha(local_destino, local_origem):
        os.makedirs(os.path.dirname(local_destino), exist_ok=True)
        shutil.copy2(local_origem, local_destino)

def enviarkey_java(nav, element_name, value):
    WebDriverWait(nav, 10).until(EC.presence_of_element_located((By.NAME, element_name)))
    script = f"document.getElementsByName('{element_name}')[0].value='{value}';" # Simula a entrada de dados via JavaScript para evitar interferência da máscara de entrada
    nav.execute_script(script)
    script = f"""
    var input = document.getElementsByName('{element_name}')[0];
    var event = new Event('input', {{ bubbles: true }});
    input.dispatchEvent(event);
    """
    nav.execute_script(script) # Dispara eventos para que o script de máscara possa processar o novo valor

def selecionar_arquivo():
    caminho_arquivo = askopenfilename(title="Selecione a Planilha COB!") # Solciita o usuario selecionar a planilha!
    return caminho_arquivo

def texto_elemento(nav,elemento,tipo):
    obj = WebDriverWait(nav, 180).until(EC.presence_of_element_located((tipo, elemento))) # Aguarda 180 segundos até o elemento carregar
    texto = obj.text
    return texto.split('-')[0]

def esperar_elementos_carregar(nav, timeout=60):
        WebDriverWait(nav, timeout).until(
            lambda driver: len(driver.find_elements(By.CLASS_NAME, "item")) > 0 or 
                           len(driver.find_elements(By.XPATH, '//div[contains(@class, "no-results-default-boxes") and contains(@class, "ng-scope") and contains(., "Sua Caixa de Entrada está vazia")]')) > 0
        )

def enviar_emails(nav, linha,click,campo):
    email = planilha.iloc[linha]['EMAILS'] # recebe e-mails da planilha (Os mesmo devem ser separados por uma '/')
    lst_email = email.split('/') # Transforma os e-mails recebidos em lista, separador '/'
    clicar_elemento(nav,click,By.XPATH) # Itens novos e-mails
    for i in range(len(lst_email)):
            acessar_iframe_default(nav) # Acessa Iframe dos e-mails
            enviarkey_elemento(nav,campo,By.ID,lst_email[i]) # Envia e-mail
            clicar_elemento_rustico(nav,'form_container',By.ID) # Clica no container para o e-mail carregar
            clicar_elemento(nav,'//*[@id="dibButtons"]/input[1]',By.XPATH) # Botão Ok
    acessar_iframe_default(nav) # Acessa Iframe dos e-mails
    clicar_elemento(nav,'cancelButtonModal',By.ID) # Botão Cancelar, para fechar janelas!

def tratar_cnpj(cnpj):
    cnpj = str(cnpj)  # Converte o CNPJ para string, caso não seja
    if '.' in cnpj or '-' in cnpj or '/' in cnpj:# Verifica se o CNPJ contém '.', '-', '/'
        cnpj_limpo = cnpj.replace('.', '').replace('/', '').replace('-', '')# Remove pontuações se existirem
    else:
        cnpj_limpo = cnpj  # Se não tiver pontuações, já está limpo
    cnpj_formatado = str(int(cnpj_limpo))# Converte para inteiro para remover zeros à esquerda, depois para string novamente
    return cnpj_formatado

def enviar_anexo(nav,linha,click,element,status,descr):
    for anexo in range(2): 
        if pd.isna(planilha.iloc[linha][f'ARQUIVO{anexo+1}']):
            pass
        else:
            clicar_elemento(nav,click,By.XPATH) # Clica no anexo para enviar arquivo
            acessar_iframe(nav)#Acesso Iframe
            enviarkey_elemento(nav,element,By.ID,fr"{caminho[:-16]}Arquivos\{planilha.iloc[linha][f'ARQUIVO{anexo+1}']}") # Envia o anexo
            while len(nav.find_elements(By.XPATH, status)) == 0: # Loop para aguardar a lista de itens carregar, se a lista não carregar a pesquisa não funciona!
                time.sleep(1)
            time.sleep(1)
            enviarkey_elemento(nav,descr,By.ID,planilha.iloc[linha][f'DESCRICAO{anexo+1}']) # Envia a descrição
            clicar_elemento(nav,'//*[@id="dibButtons"]/input[1]',By.XPATH) 
            nav.switch_to.default_content()#Volta para o inicio

def opcoes_pagamento(nav,selec,seta):
    for i in range(2):#Loop para selecionar as opções de pagamento
        nav.find_element(By.XPATH, selec).click()
        nav.find_element(By.ID, seta).click()  

def clicar_porcentagem(nav,contador,linha):
    for i in range(contador): # Baseado na soma do Contador clica nos itens
        clicar_elemento_rustico(nav,f'//*[@id="{i}"]/td[2]',By.XPATH) # Clica no Item baseado nos indices (No fusion o indice 0 conta!)!!
        acessar_iframe_default(nav) # Acessa Iframe primario
        clicar_elemento(nav,'action.save',By.NAME) # Clica para salvar.
        acessar_iframe_default(nav) # Acessa Iframe primario
    enviarkey_elemento(nav,'id_txt_dadosDaCobranca__dadosDoFaturamentoVariavel__numeroContratoProtheus__',By.ID,str(int(planilha.iloc[linha]['NUMERO DO CONTRATO']))) # Envia o numero de contrato
    clicar_elemento(nav,'//*[@id="ui-id-10"]/li',By.XPATH) # Clica no numero de contrato
    clicar_elemento(nav,'action.save',By.NAME) # Clica para salvar.
    nav.switch_to.default_content()#Volta para o inicio

def dados_rateio(nav,linha,cod_filial,cod_uo):
    clicar_elemento(nav,'//*[@id="menu_bar_FINFFCobFaturamentoVariavelCentroDeResultadosXFilialXValor"]/li[1]',By.XPATH) # Clica para abrir campo de produtos
    acessar_iframe_default(nav) # Acessa Iframe da Pesquisa de produtos
    opcoes_pagamento(nav,'//*[@id="mul_dadosDaCobranca__dadosDoFaturamentoVariavel__dadosDoRateio__formaDeEntradaDosRecursos_ori"]/option[1]','move_this_right_mul_dadosDaCobranca__dadosDoFaturamentoVariavel__dadosDoRateio__formaDeEntradaDosRecursos')#Loop para selecionar as opções de pagamento
    clicar_elemento(nav,'id_dadosDaCobranca__dadosDoFaturamentoVariavel__dadosDoRateio__UOCRProtheus___anchor',By.ID) # Clica para abrir campo de pesquisa
    acessar_iframe_default(nav) # Acessa Iframe da Pesquisa
    clicar_elemento(nav,'//*[@id="menu_bar_EXTERNOProtheusAmarracaoContabil"]/li',By.XPATH) # Clica para abrir filtro
    acessar_iframe_default(nav) # Acessa Iframe do Filtro
    enviarkey_elemento(nav,'var_codclvlr__',By.NAME,str(planilha.iloc[linha]['CLASSE DE VALOR'])) # Envia Classa de valor Cliente
    enviarkey_elemento(nav,'var_codfilialprotheus__',By.NAME,cod_filial) # Envia COD FILIAL - PADRÃO
    enviarkey_elemento(nav,'var_coduo__',By.NAME,cod_uo) # Envia COD UO - PADRÃO

def iniciar_navegador():
    servico = Service(ChromeDriverManager().install())# Start no Navegador Chrome
    options = webdriver.ChromeOptions()
    options.add_experimental_option("detach", True)# Para o mesmo não fechar apos execução
    navegador = webdriver.Chrome(options=options, service=servico)# Executa o navegador
    navegador.get('https://fusion.fiemg.com.br/fusion/portal')
    navegador.maximize_window()# Maximiza a janela do navegador
    return navegador
    
def variavel_novo(nav,linha):
        enviarkey_elemento(nav,'id_tipoDeMedicao__',By.ID,'Variavel')# Tipo de medição
        enviarkey_elemento(nav,'var_dadosDaCobranca__APeriodicidadeDoFaturamentoEMensal__',By.ID,'Sim')# Periodicidade
        enviarkey_elemento(nav,'var_dadosDaCobranca__dadosParaHistorico__HouvePrestacaoDeServicos__',By.ID,'Sim')# Prestação de serviço
        enviarkey_elemento(nav,'id_txt_dadosDaCobranca__dadosParaHistorico__numeroContratoProtheus__',By.ID,str(int(planilha.iloc[linha]['NUMERO DO CONTRATO'])))# Numero do Contrato
        clicar_elemento(nav,'//*[@id="ui-id-3"]/li',By.XPATH) # Clica no Numero do contrato
        enviarkey_elemento(nav,'var_dadosDaCobranca__dadosParaHistorico__existeGasOuOS__',By.ID,'Não')# Existe GAS ou OS
        enviarkey_elemento(nav,'var_dadosDaCobranca__dadosParaHistorico__parcelaContrato__',By.ID,'0')# Informa Parcela
        enviarkey_elemento(nav,'var_dadosDaCobranca__dadosParaHistorico__inicioPrestacao__',By.ID,primeiro_dia)# Data Inicio
        enviarkey_elemento(nav,'var_dadosDaCobranca__dadosParaHistorico__finPrestacao__',By.ID,ultimo_dia)# Data Fim
        enviarkey_elemento(nav,'id_dadosDaCobranca__dadosParaHistorico__diaLimiteNFCliente__',By.ID,str(int(planilha.iloc[linha]['DIA LIMITE'])))# Dia Limite
        enviarkey_elemento(nav,'var_dadosDaCobranca__cobrancaRelacionadaComConvenio__',By.ID,'Não')# Convenio

cod_filial = '01MG0014' # Codigo Filial - Padrão
cod_uo = '10310' # Codigo UO - Padrão

while True:
    print('-----------Automação COB-----------/n')
    print("\nMenu de Escolhas:")
    print("1 - Medição Variavel")
    print("2 - Criar Novos COB's")
    print("0 - Sair")
    escolha = input("Escolha uma opção: ")
    try:
        escolha = int(escolha)
        if escolha == 1:
            usuario = input('Insira o usuario do Fusion: ')
            senha = getpass('Insira a senha do Fusion: ')
            caminho = selecionar_arquivo()
            planilha = pd.read_excel(caminho,'Medição') # Carrega a Planilha
            destino = os.path.dirname(caminho)# Pega o caminho da pasta
            planilha_destino = destino + r'/Historico.xlsx' # Caminho do Historico
            local_destino = r'C:\Temp\Historico.xlsx'
            copiar_para_planilha(local_destino, planilha_destino)
            navegador = iniciar_navegador()
            enviarkey_elemento(navegador,'user',By.ID,usuario)# Login
            enviarkey_elemento(navegador,'pass',By.ID,senha)# Senha
            clicar_elemento(navegador,'btnLogin',By.ID) # Clica no botão de Login
            acessar_iframe_default(navegador)# Acessa o Iframe
            for linha in range(len(planilha)):
                enviarkey_elemento(navegador,'searchBarProcessQuery',By.ID,planilha.iloc[linha]['COB'])#Envio do COB
                esperar_elementos_carregar(navegador)
                clicar_elemento_rustico(navegador,'//*[@id="page-content-wrapper"]/div/div/div[1]/div[1]/nav/div/form/div/div/span/button',By.XPATH) # Clica no botão de pesquisa inicial
                aba_orignal = navegador.window_handles[0] # Identifica Aba Primaria
                clicar_elemento_rustico(navegador, 'header', By.CLASS_NAME) # Clica no COB pesquisado
                WebDriverWait(navegador, 10).until(lambda d: len(d.window_handles) > 1)
                nova_aba = navegador.window_handles[1]# Identifica nova aba apos iniciar Cobrança
                navegador.switch_to.window(nova_aba) # Troca para nova Aba
                data = planilha.iloc[linha]['DATA_DESCRIÇÃO'] # Pega data de Descrição
                date = datetime.strptime(data.strftime('%d/%m/%Y'), '%d/%m/%Y') # Transforma data em string
                primeiro_dia, ultimo_dia = primeiro_e_ultimo_dia_do_mes(date.year, date.month) # Pega o mês e dia
                if pd.isna(planilha.iloc[linha]['DATA_DE_VENCIMENTO']):
                    data_venc = date
                else:
                    data_venc = planilha.iloc[linha]['DATA_DE_VENCIMENTO'] # Pega data de Vencimento
                # ---------------------- Esta Parte se refere ao COB sem Rateio ------------------------
                for sem_rateio in range(2):
                    if pd.isna(planilha.iloc[linha][f'CR-SR{sem_rateio+1}']): # Verifica se o campo está vazio
                        pass
                    else:
                        clicar_elemento(navegador,'createitem',By.ID)# Clica para criar novo Item
                        acessar_iframe(navegador)# Acessa o Iframe
                        enviarkey_elemento(navegador,'id_txt_dadosDaCobranca__dadosDoFaturamentoVariavel__dadosDoCliente__',By.ID,tratar_cnpj(planilha.iloc[linha]['CNPJ'])) # Envia CNPJ
                        clicar_elemento(navegador,'ui-id-11',By.ID) # Clica no CNPJ informado
                        if sem_rateio == 0: # Difere o primeiro produto do segundo
                            enviarkey_elemento(navegador,'var_dadosDaCobranca__dadosDoFaturamentoVariavel__descricaoServico__',By.NAME,f'COBRANÇA SESI VIVA+: AEP,PGR,PCMSO,LTCAT \nPERÍODO: {primeiro_dia} a {ultimo_dia}.') # Envia Descrição
                        else:
                            enviarkey_elemento(navegador,'var_dadosDaCobranca__dadosDoFaturamentoVariavel__descricaoServico__',By.NAME,f'COBRANÇA CONSULTAS E EXAMES COMPLEMENTARES. \nPERÍODO: {primeiro_dia} a {ultimo_dia}.') # Envia Descrição
                        enviarkey_elemento(navegador,'var_dadosDaCobranca__dadosDoFaturamentoVariavel__rateio__', By.NAME,'Não')# Envia não ao campo de rateio
                        clicar_elemento(navegador,'id_dadosDaCobranca__dadosDoFaturamentoVariavel__dadosDeCobranca__UOCRProtheus___anchor',By.ID)# Clica na pesquisa de produto
                        acessar_iframe_default(navegador) # Acessa Iframe da Pesquisa
                        clicar_elemento(navegador,'vfilter',By.ID) # Clica no Filtro
                        acessar_iframe_default(navegador) # Acessa Iframe do Filtro
                        enviarkey_elemento(navegador,'var_codclvlr__',By.NAME,str(int(planilha.iloc[linha]['CLASSE DE VALOR']))) # Envia Classe de valor Cliente
                        enviarkey_elemento(navegador,'var_codfilialprotheus__',By.NAME,cod_filial) # Envia COD FILIAL - PADRÃO
                        enviarkey_elemento(navegador,'var_coduo__',By.NAME,cod_uo) # Envia COD UO - PADRÃO
                        enviarkey_elemento(navegador,'var_codccusto__',By.NAME,str(int(planilha.iloc[linha][f'CR-SR{sem_rateio+1}']))) # Envia COD PRODUTO
                        clicar_elemento(navegador,'searchbutton',By.ID) # Clica na Pesquisa
                        acessar_iframe_default(navegador) # Acessa Iframe da Pesquisa
                        clicar_elemento(navegador,'tooltip0',By.ID)
                        acessar_iframe_default(navegador) # Acessa Iframe primario
                        clicar_elemento(navegador,'createitem',By.ID) # Clica para adicionar Valor
                        acessar_iframe_default(navegador) # Acessa Iframe de valor
                        opcoes_pagamento(navegador,'//*[@id="mul_dadosDaCobranca__dadosDoFaturamentoVariavel__dataVencimentoValorCobranca__formaDeEntradaDosRecursos_ori"]/option[1]','move_this_right_mul_dadosDaCobranca__dadosDoFaturamentoVariavel__dataVencimentoValorCobranca__formaDeEntradaDosRecursos')#Loop para selecionar as opções de pagamento
                        enviarkey_elemento(navegador,'var_dadosDaCobranca__dadosDoFaturamentoVariavel__dataVencimentoValorCobranca__data__',By.NAME,data_venc.strftime('%d/%m/%Y')) # Envia data da cobrança
                        enviarkey_java(navegador,'var_dadosDaCobranca__dadosDoFaturamentoVariavel__dataVencimentoValorCobranca__valor__',planilha.iloc[linha][f'VALORSR{sem_rateio+1}']) # Envia Valor
                        clicar_elemento(navegador,'action.save',By.NAME) # Clica para salvar.
                        acessar_iframe_default(navegador) # Acessa Iframe primario
                        enviarkey_elemento(navegador,'id_txt_dadosDaCobranca__dadosDoFaturamentoVariavel__numeroContratoProtheus__',By.ID,str(int(planilha.iloc[linha]['NUMERO DO CONTRATO']))) # Envia o numero de contrato
                        clicar_elemento(navegador,'//*[@id="ui-id-10"]/li',By.XPATH) # Clica no numero de contrato
                        clicar_elemento(navegador,'action.save',By.NAME) # Clica para salvar.
                        navegador.switch_to.default_content()#Volta para o inicio

                # ---------------------- Esta Parte se refere ao COB com Rateio ------------------------

                contador = 0 # Contador utilizado para clicar nos rateios no processo Final!
                contador_1 = 0 # Contador utilizado para clicar nos rateios no processo Final!
                contador_2 = 0 # Contador utilizado para clicar nos rateios no processo Final!
                if pd.isna(planilha.iloc[linha]['CRR1']): # Verifica se o primeiro item está vazio, se o mesmo estiver vazio, todo o loop é pulado!
                    pass

                elif pd.isna(planilha.iloc[linha]['QTD RATEIO']) or int(planilha.iloc[linha]['QTD RATEIO']) == 1 : # Caso não esteja vazio é iniciado o processo de Rateio e a QTD seja um executa todos os rateios em um unico processo
                    clicar_elemento(navegador,'createitem',By.ID)# Clica para criar novo Item
                    acessar_iframe(navegador)# Acessa o Iframe
                    enviarkey_elemento(navegador,'id_txt_dadosDaCobranca__dadosDoFaturamentoVariavel__dadosDoCliente__',By.ID,tratar_cnpj(planilha.iloc[linha]['CNPJ'])) # Envia CNPJ
                    clicar_elemento(navegador,'ui-id-11',By.ID) # Clica no CNPJ informado
                    enviarkey_elemento(navegador,'var_dadosDaCobranca__dadosDoFaturamentoVariavel__descricaoServico__',By.NAME,f'COBRANÇA CONSULTAS E EXAMES COMPLEMENTARES. \nPERÍODO: {primeiro_dia} a {ultimo_dia}.') # Envia Descrição
                    enviarkey_elemento(navegador,'var_dadosDaCobranca__dadosDoFaturamentoVariavel__rateio__', By.NAME,'Sim')# Envia sim ao campo de rateio
                    #Loop para a quantidade de Itens
                    for com_rateio in range(4): # Loop para verificar todos os itens (Total 4) com rateio na planilha!!
                        if pd.isna(planilha.iloc[linha][f'CRR{com_rateio+1}']): # Loop para verificar se o Item está vazio!!
                            pass # Pula o item vazio
                        else:
                            dados_rateio(navegador,linha,cod_filial,cod_uo)
                            enviarkey_elemento(navegador,'var_codccusto__',By.NAME,str(int(planilha.iloc[linha][f'CRR{com_rateio+1}']))) # Envia COD PRODUTO
                            clicar_elemento(navegador,'searchbutton',By.ID) # Clica na Pesquisa
                            acessar_iframe_default(navegador) # Acessa Iframe da Pesquisa
                            clicar_elemento(navegador,'tooltip0',By.ID) # Clica no item filtrado
                            acessar_iframe_default(navegador) # Acessa Iframe primario3
                            enviarkey_java(navegador,'var_dadosDaCobranca__dadosDoFaturamentoVariavel__dadosDoRateio__valor__',planilha.iloc[linha][f'VALOR{com_rateio+1}']) # Envia Valor
                            clicar_elemento(navegador,'action.save',By.NAME) # Clica para salvar.
                            acessar_iframe_default(navegador) # Acessa Iframe primario
                            contador += 1 # Soma 1 a quantidade de contador, será utiizado para clicar no loop Contador!
                    clicar_porcentagem(navegador,contador,linha) # Baseado na soma do Contador clica nos itens

                elif int(planilha.iloc[linha]['QTD RATEIO']) == 2: # Ira rodar o processo de sem rateio duas vezes uma para a coluna CRR1 e 2 e ou para CRR3 e 4
                    # Processo para coluna 1 e 2
                    clicar_elemento(navegador,'createitem',By.ID)# Clica para criar novo Item
                    acessar_iframe(navegador)# Acessa o Iframe
                    enviarkey_elemento(navegador,'id_txt_dadosDaCobranca__dadosDoFaturamentoVariavel__dadosDoCliente__',By.ID,tratar_cnpj(planilha.iloc[linha]['CNPJ'])) # Envia CNPJ
                    clicar_elemento(navegador,'ui-id-11',By.ID) # Clica no CNPJ informado
                    enviarkey_elemento(navegador,'var_dadosDaCobranca__dadosDoFaturamentoVariavel__descricaoServico__',By.NAME,f'COBRANÇA CONSULTAS E EXAMES COMPLEMENTARES. \nPERÍODO: {primeiro_dia} a {ultimo_dia}.') # Envia Descrição
                    enviarkey_elemento(navegador,'var_dadosDaCobranca__dadosDoFaturamentoVariavel__rateio__', By.NAME,'Sim')# Envia sim ao campo de rateio
                    #Loop para a quantidade de Itens
                    for com_rateio in range(2): # Loop para verificar todos os itens (Total 4) com rateio na planilha!!
                        if pd.isna(planilha.iloc[linha][f'CRR{com_rateio+1}']): # Loop para verificar se o Item está vazio!!
                            pass # Pula o item vazio
                        else:
                            dados_rateio(navegador,linha,cod_filial,cod_uo)
                            enviarkey_elemento(navegador,'var_codccusto__',By.NAME,str(int(planilha.iloc[linha][f'CRR{com_rateio+1}']))) # Envia COD PRODUTO
                            clicar_elemento(navegador,'searchbutton',By.ID) # Clica na Pesquisa
                            acessar_iframe_default(navegador) # Acessa Iframe da Pesquisa
                            clicar_elemento(navegador,'tooltip0',By.ID) # Clica no item filtrado
                            acessar_iframe_default(navegador) # Acessa Iframe primario3
                            enviarkey_java(navegador,'var_dadosDaCobranca__dadosDoFaturamentoVariavel__dadosDoRateio__valor__',planilha.iloc[linha][f'VALOR{com_rateio+1}']) # Envia Valor
                            clicar_elemento(navegador,'action.save',By.NAME) # Clica para salvar.
                            acessar_iframe_default(navegador) # Acessa Iframe primario
                            contador_1 += 1 # Soma 1 a quantidade de contador, será utiizado para clicar no loop Contador!
                    clicar_porcentagem(navegador,contador_1,linha) # Baseado na soma do Contador clica nos itens

                    # Processo para coluna 3 e 4
                    clicar_elemento(navegador,'createitem',By.ID)# Clica para criar novo Item
                    acessar_iframe(navegador)# Acessa o Iframe
                    enviarkey_elemento(navegador,'id_txt_dadosDaCobranca__dadosDoFaturamentoVariavel__dadosDoCliente__',By.ID,tratar_cnpj(planilha.iloc[linha]['CNPJ'])) # Envia CNPJ
                    clicar_elemento(navegador,'ui-id-11',By.ID) # Clica no CNPJ informado
                    enviarkey_elemento(navegador,'var_dadosDaCobranca__dadosDoFaturamentoVariavel__descricaoServico__',By.NAME,f'IN LOCO COBRANÇA CONSULTAS E EXAMES COMPLEMENTARES. \nPERÍODO: {primeiro_dia} a {ultimo_dia}.') # Envia Descrição
                    enviarkey_elemento(navegador,'var_dadosDaCobranca__dadosDoFaturamentoVariavel__rateio__', By.NAME,'Sim')# Envia sim ao campo de rateio
                    #Loop para a quantidade de Itens
                    for com_rateio in range(2): # Loop para verificar todos os itens (Total 4) com rateio na planilha!!
                        if pd.isna(planilha.iloc[linha][f'CRR{com_rateio+3}']): # Loop para verificar se o Item está vazio!!
                            pass # Pula o item vazio
                        else:
                            dados_rateio(navegador,linha,cod_filial,cod_uo)
                            enviarkey_elemento(navegador,'var_codccusto__',By.NAME,str(int(planilha.iloc[linha][f'CRR{com_rateio+3}']))) # Envia COD PRODUTO
                            clicar_elemento(navegador,'searchbutton',By.ID) # Clica na Pesquisa
                            acessar_iframe_default(navegador) # Acessa Iframe da Pesquisa
                            clicar_elemento(navegador,'tooltip0',By.ID) # Clica no item filtrado
                            acessar_iframe_default(navegador) # Acessa Iframe primario3
                            enviarkey_java(navegador,'var_dadosDaCobranca__dadosDoFaturamentoVariavel__dadosDoRateio__valor__',planilha.iloc[linha][f'VALOR{com_rateio+3}']) # Envia Valor
                            clicar_elemento(navegador,'action.save',By.NAME) # Clica para salvar.
                            acessar_iframe_default(navegador) # Acessa Iframe primario
                            contador_2 += 1 # Soma 1 a quantidade de contador, será utiizado para clicar no loop Contador!
                    clicar_porcentagem(navegador,contador_2,linha) # Baseado na soma do Contador clica nos itens

                # ---------------------- Esta Parte se refere aos Anexos ------------------------
                enviar_anexo(navegador,linha,'//*[@id="menu_bar_genericoHistoricoAtendimento"]/li[1]','var_dadosDaCobranca__historico__anexo__','//*[@id="progress-complete-var_dadosDaCobranca__historico__anexo__"]/span','var_dadosDaCobranca__historico__registro__') # Envia Anexos
                if len(navegador.find_elements(By.ID, 'id_dadosDaCobranca__acao__')) >= 1: # Verifica se o campo existe
                    enviarkey_elemento(navegador,'id_dadosDaCobranca__acao__',By.ID,'Solicitar Nova Medição')
                input('Confirma o lançamento!!!')
                clicar_elemento(navegador,'action.send',By.NAME)
                while len(navegador.find_elements(By.CLASS_NAME, 'alert')) == 0: # Loop para aguardar o alerta carregar!
                    time.sleep(1)
                time.sleep(1)
                copiar_linha_ativa(planilha, local_destino, 'Medição', linha)
                navegador.close() # Fecha a aba apos Alerta Carregar!!
                navegador.switch_to.window(aba_orignal)
                time.sleep(1)
                acessar_iframe_default(navegador)
                clicar_elemento_rustico(navegador,'clear-input-filter',By.CLASS_NAME)#Limpa o campo de Pesquisa
            copiar_para_planilha(planilha_destino,local_destino)

        elif escolha == 2:
            usuario = input('Insira o usuario do Fusion: ')
            senha = getpass('Insira a senha do Fusion: ')
            caminho = selecionar_arquivo()
            planilha = pd.read_excel(caminho,'Novo') # Carrega a Planilha
            destino = os.path.dirname(caminho)
            planilha_destino = destino + r'/Historico.xlsx'
            local_destino = r'C:\Temp\Historico.xlsx'
            copiar_para_planilha(local_destino, planilha_destino)
            navegador = iniciar_navegador() #Inicia o Navegador
            enviarkey_elemento(navegador,'user',By.ID,usuario)# Login
            enviarkey_elemento(navegador,'pass',By.ID,senha)# Senha
            clicar_elemento(navegador,'btnLogin',By.ID) # Clica no botão de Login
            acessar_iframe_default(navegador)# Acessa o Iframe
            aba_orignal = navegador.window_handles[0] # Identifica Aba Primaria

            for linha in range(len(planilha)):
                clicar_elemento(navegador,'btnStartProcess',By.ID) # Iniciar novo processo
                clicar_elemento(navegador,'//*[@id="page-content-wrapper"]/div/div/div[1]/div[1]/nav/div/div/div/ul/li[3]/ul/li[5]/a/div/span[1]',By.XPATH) # Iniciar nova Cobrança
                WebDriverWait(navegador, 10).until(lambda d: len(d.window_handles) > 1)
                nova_aba = navegador.window_handles[1]# Identifica nova aba apos iniciar nova Cobrança
                navegador.switch_to.window(nova_aba) # Troca para nova Aba
                enviarkey_elemento(navegador,'id_informeNucleo__',By.ID,'Núcleo de Faturamento')# Envia nucleo - Padrão
                enviarkey_elemento(navegador,'id_tipoSolicitacao__',By.ID,'Solicitação de cobrança (FG-176)')# Solicitação de cobrança - Padrão
                enviarkey_elemento(navegador,'id_plataformaGestaoDaVenda__',By.ID,'Protheus')# Plataforma - Padrão

                if planilha.iloc[linha]['TIPO'].lower() == "variavel" and planilha.iloc[linha]['RATEIO'].lower() == "sim":

                    data = planilha.iloc[linha]['DESCRICAO']
                    date = datetime.strptime(data.strftime('%d/%m/%Y'), '%d/%m/%Y') # Transforma data em string
                    primeiro_dia, ultimo_dia = primeiro_e_ultimo_dia_do_mes(date.year, date.month) # Pega o mês e dia
                    data_str = planilha.iloc[linha]['DATA']
                    data_obj = datetime.strptime(data_str.strftime('%d/%m/%Y'), '%d/%m/%Y')
                    nome_cob = texto_elemento(navegador,'headerTitle',By.ID)
                    variavel_novo(navegador,linha) # Carrega dados de preenchimento
                    if pd.isna(planilha.iloc[linha][f'CR1']): # Loop para verificar se o Item está vazio!!
                        pass
                    else:
                        clicar_elemento(navegador,'//*[@id="createitem"]',By.XPATH) # Clica no Novo Item
                        acessar_iframe(navegador)# Acessa o Iframe
                        enviarkey_elemento(navegador,'id_txt_dadosDaCobranca__dadosDoFaturamentoVariavel__dadosDoCliente__',By.ID,tratar_cnpj(planilha.iloc[linha]['CNPJ'])) # Envia CNPJ
                        clicar_elemento(navegador,'ui-id-11',By.ID) # Clica no CNPJ informado
                        enviarkey_elemento(navegador,'var_dadosDaCobranca__dadosDoFaturamentoVariavel__descricaoServico__',By.NAME,f'COBRANÇA SESI VIVA+: AEP,PGR,PCMSO,LTCAT \nPERÍODO: {primeiro_dia} a {ultimo_dia}.') # Envia Descrição
                        enviarkey_elemento(navegador,'var_dadosDaCobranca__dadosDoFaturamentoVariavel__rateio__', By.NAME,'Não')# Envia não ao campo de rateio
                        clicar_elemento(navegador,'id_dadosDaCobranca__dadosDoFaturamentoVariavel__dadosDeCobranca__UOCRProtheus___anchor',By.ID) # Clica na pesquisa
                        acessar_iframe_default(navegador) # Acessa Iframe da Pesquisa
                        clicar_elemento(navegador,'vfilter',By.ID) # Clica no Filtro
                        acessar_iframe_default(navegador) # Acessa Iframe do Filtro
                        enviarkey_elemento(navegador,'var_codclvlr__',By.NAME,int(planilha.iloc[linha]['CLASSE DE VALOR'])) # Envia Classa de valor Cliente
                        enviarkey_elemento(navegador,'var_codfilialprotheus__',By.NAME,cod_filial) # Envia COD FILIAL - PADRÃO
                        enviarkey_elemento(navegador,'var_coduo__',By.NAME,cod_uo) # Envia COD UO - PADRÃO
                        enviarkey_elemento(navegador,'var_codccusto__',By.NAME,int(planilha.iloc[linha]['CR1'])) # Envia COD PRODUTO
                        clicar_elemento(navegador,'searchbutton',By.ID) # Clica na Pesquisa
                        acessar_iframe_default(navegador) # Acessa Iframe da Pesquisa
                        clicar_elemento(navegador,'tooltip0',By.ID)
                        acessar_iframe_default(navegador) # Acessa Iframe primario
                        clicar_elemento(navegador,'createitem',By.ID) # Clica para adicionar Valor
                        acessar_iframe_default(navegador) # Acessa Iframe de valor
                        opcoes_pagamento(navegador,'//*[@id="mul_dadosDaCobranca__dadosDoFaturamentoVariavel__dataVencimentoValorCobranca__formaDeEntradaDosRecursos_ori"]/option[1]','move_this_right_mul_dadosDaCobranca__dadosDoFaturamentoVariavel__dataVencimentoValorCobranca__formaDeEntradaDosRecursos')#Loop para selecionar as opções de pagamento     
                        enviarkey_elemento(navegador,'var_dadosDaCobranca__dadosDoFaturamentoVariavel__dataVencimentoValorCobranca__data__',By.ID,data_obj.strftime('%d/%m/%Y')) # Envia data da cobrança
                        enviarkey_java(navegador,'var_dadosDaCobranca__dadosDoFaturamentoVariavel__dataVencimentoValorCobranca__valor__',planilha.iloc[linha]['VALOR1'])
                        clicar_elemento(navegador,'action.save',By.NAME) # Clica para salvar.
                        acessar_iframe_default(navegador) # Acessa Iframe primario
                    enviarkey_elemento(navegador,'id_txt_dadosDaCobranca__dadosDoFaturamentoVariavel__numeroContratoProtheus__',By.ID,int(planilha.iloc[linha]['NUMERO DO CONTRATO'])) # Envia o numero de contrato
                    clicar_elemento(navegador,'//*[@id="ui-id-10"]/li',By.XPATH) # Clica no numero de contrato
                    clicar_elemento(navegador,'//*[@id="dibButtons"]/input[1]',By.XPATH) # Clica no numero de contrato
                    navegador.switch_to.default_content()#Volta para o inicio
                    #------------------------Rateio--------------------------------------------------
                    clicar_elemento(navegador,'//*[@id="createitem"]',By.XPATH) # Clica no Novo
                    acessar_iframe(navegador)# Acessa o Iframe
                    enviarkey_elemento(navegador,'id_txt_dadosDaCobranca__dadosDoFaturamentoVariavel__dadosDoCliente__',By.ID,tratar_cnpj(planilha.iloc[linha]['CNPJ'])) # Envia CNPJ
                    clicar_elemento(navegador,'ui-id-11',By.ID) # Clica no CNPJ informado
                    enviarkey_elemento(navegador,'var_dadosDaCobranca__dadosDoFaturamentoVariavel__descricaoServico__',By.NAME,f'COBRANÇA CONSULTAS E EXAMES COMPLEMENTARES. \nPERÍODO: {primeiro_dia} a {ultimo_dia}.') # Envia Descrição
                    enviarkey_elemento(navegador,'var_dadosDaCobranca__dadosDoFaturamentoVariavel__rateio__', By.NAME,'SIM')# Envia Sim ao campo de rateio

                    contador = 0 # Contador utilizado para clicar nos rateios no processo Final!
                    #Loop para a quantidade de Itens
                    for com_rateio in range(5): # Loop para verificar todos os itens (Total 6) com rateio na planilha!!
                        if pd.isna(planilha.iloc[linha][f'CR{com_rateio+2}']): # Loop para verificar se o Item está vazio!!
                            pass # Pula o item vazio
                        else:
                            dados_rateio(navegador,linha,cod_filial,cod_uo)
                            enviarkey_elemento(navegador,'var_codccusto__',By.NAME,str(int(planilha.iloc[linha][f'CR{com_rateio+2}']))) # Envia COD PRODUTO
                            clicar_elemento(navegador,'searchbutton',By.ID) # Clica na Pesquisa
                            acessar_iframe_default(navegador) # Acessa Iframe da Pesquisa
                            clicar_elemento(navegador,'tooltip0',By.ID) # Clica no item filtrado
                            acessar_iframe_default(navegador) # Acessa Iframe primario3
                            elemento2 = WebDriverWait(navegador, 10).until(EC.presence_of_element_located((By.NAME, 'var_dadosDaCobranca__dadosDoFaturamentoVariavel__dadosDoRateio__valor__')))
                            script_valor_cr = f"document.getElementsByName('var_dadosDaCobranca__dadosDoFaturamentoVariavel__dadosDoRateio__valor__')[0].value='{planilha.iloc[linha][f'VALOR{com_rateio+2}']}';"
                            navegador.execute_script(script_valor_cr)
                            clicar_elemento(navegador,'action.save',By.NAME) # Clica para salvar.
                            acessar_iframe_default(navegador) # Acessa Iframe primario
                            contador += 1 # Soma 1 a quantidade de contador, será utiizado para clicar no loop Contador!

                    for i in range(contador): # Baseado na soma do Contador clica nos itens
                        clicar_elemento_rustico(navegador,f'//*[@id="{i}"]/td[2]',By.XPATH) # Clica no Item baseado nos indices (No fusion o indice 0 conta!)!!
                        acessar_iframe_default(navegador) # Acessa Iframe primario
                        clicar_elemento(navegador,'action.save',By.NAME) # Clica para salvar.
                        acessar_iframe_default(navegador) # Acessa Iframe primario

                    enviarkey_elemento(navegador,'id_txt_dadosDaCobranca__dadosDoFaturamentoVariavel__numeroContratoProtheus__',By.ID,str(int(planilha.iloc[linha]['NUMERO DO CONTRATO']))) # Envia o numero de contrato
                    clicar_elemento(navegador,'//*[@id="ui-id-10"]/li',By.XPATH) # Clica no numero de contrato
                    clicar_elemento(navegador,'action.save',By.NAME) # Clica para salvar.
                    navegador.switch_to.default_content()#Volta para o inicio
                    
                    enviar_anexo(navegador,linha,'//*[@id="menu_bar_genericoHistoricoAtendimento"]/li[1]','var_dadosDaCobranca__historico__anexo__','//*[@id="progress-complete-var_dadosDaCobranca__historico__anexo__"]/span','var_dadosDaCobranca__historico__registro__')#Envia Anexos
                    enviar_emails(navegador,linha,"//li[@onclick=\"activeDeactiveObjMenu2(this);javascript: ellist_emailClienteFP__.addNewItem('CreateItens', true);\"]//a[@id='createitens']",'var_emailClienteFP__Email__') # Envia e-mails
                    input('Confirma o lançamento!!!')
                    navegador.switch_to.default_content()
                    clicar_elemento(navegador,'action.send',By.NAME)
                    while len(navegador.find_elements(By.CLASS_NAME, 'alert')) == 0: # Loop para aguardar o alerta carregar!
                        time.sleep(1)
                    time.sleep(1)
                    copiar_linha_ativa(planilha, local_destino, 'Novo', linha,nome_cob)
                    navegador.close() # Fecha a aba apos Alerta Carregar!!
                    navegador.switch_to.window(aba_orignal)
                    time.sleep(1)
                    acessar_iframe_default(navegador)

                elif planilha.iloc[linha]['TIPO'].lower() == "variavel" and planilha.iloc[linha]['RATEIO'].lower() == "não":

                    data = planilha.iloc[linha]['DESCRICAO']
                    date = datetime.strptime(data.strftime('%d/%m/%Y'), '%d/%m/%Y') # Transforma data em string
                    primeiro_dia, ultimo_dia = primeiro_e_ultimo_dia_do_mes(date.year, date.month) # Pega o mês e dia
                    data_str = planilha.iloc[linha]['DATA']
                    data_obj = datetime.strptime(data_str.strftime('%d/%m/%Y'), '%d/%m/%Y')
                    nome_cob = texto_elemento(navegador,'headerTitle',By.ID)
                    variavel_novo(navegador,linha) # Carrega dados de preenchimento
                    clicar_elemento(navegador,'//*[@id="createitem"]',By.XPATH) #  Clica no Novo Item
                    acessar_iframe(navegador)# Acessa o Iframe
                    enviarkey_elemento(navegador,'id_txt_dadosDaCobranca__dadosDoFaturamentoVariavel__dadosDoCliente__',By.ID,tratar_cnpj(planilha.iloc[linha]['CNPJ'])) # Envia CNPJ
                    clicar_elemento(navegador,'ui-id-11',By.ID) # Clica no CNPJ informado
                    enviarkey_elemento(navegador,'var_dadosDaCobranca__dadosDoFaturamentoVariavel__descricaoServico__',By.NAME,f'COBRANÇA SESI VIVA+: AEP,PGR,PCMSO,LTCAT \nPERÍODO: {primeiro_dia} a {ultimo_dia}.') # Envia Descrição
                    enviarkey_elemento(navegador,'var_dadosDaCobranca__dadosDoFaturamentoVariavel__rateio__', By.NAME,'Não')# Envia não ao campo de rateio
                    for sem_rateio in range(2):
                        if pd.isna(planilha.iloc[linha][f'CR{sem_rateio+1}']):
                            pass
                        else:
                            clicar_elemento(navegador,'id_dadosDaCobranca__dadosDoFaturamentoVariavel__dadosDeCobranca__UOCRProtheus___anchor',By.ID)# Clica novo
                            acessar_iframe_default(navegador) # Acessa Iframe da Pesquisa
                            clicar_elemento(navegador,'vfilter',By.ID) # Clica no Filtro
                            acessar_iframe_default(navegador) # Acessa Iframe do Filtro
                            enviarkey_elemento(navegador,'var_codclvlr__',By.NAME,int(planilha.iloc[linha]['CLASSE DE VALOR'])) # Envia Classa de valor Cliente
                            enviarkey_elemento(navegador,'var_codfilialprotheus__',By.NAME,cod_filial) # Envia COD FILIAL - PADRÃO
                            enviarkey_elemento(navegador,'var_coduo__',By.NAME,cod_uo) # Envia COD UO - PADRÃO
                            enviarkey_elemento(navegador,'var_codccusto__',By.NAME,int(planilha.iloc[linha][f'CR{sem_rateio+1}'])) # Envia COD PRODUTO
                            clicar_elemento(navegador,'searchbutton',By.ID) # Clica na Pesquisa
                            acessar_iframe_default(navegador) # Acessa Iframe da Pesquisa
                            clicar_elemento(navegador,'tooltip0',By.ID)
                            acessar_iframe_default(navegador) # Acessa Iframe primario
                            clicar_elemento(navegador,'createitem',By.ID) # Clica para adicionar Valor
                            acessar_iframe_default(navegador) # Acessa Iframe de valor
                            opcoes_pagamento(navegador,'//*[@id="mul_dadosDaCobranca__dadosDoFaturamentoVariavel__dataVencimentoValorCobranca__formaDeEntradaDosRecursos_ori"]/option[1]','move_this_right_mul_dadosDaCobranca__dadosDoFaturamentoVariavel__dataVencimentoValorCobranca__formaDeEntradaDosRecursos')#Loop para selecionar as opções de pagamento 
                            enviarkey_elemento(navegador,'var_dadosDaCobranca__dadosDoFaturamentoVariavel__dataVencimentoValorCobranca__data__',By.ID,data_obj.strftime('%d/%m/%Y')) # Envia data da cobrança
                            enviarkey_java(navegador,'var_dadosDaCobranca__dadosDoFaturamentoVariavel__dataVencimentoValorCobranca__valor__',planilha.iloc[linha][f'VALOR{sem_rateio+1}'])
                            clicar_elemento(navegador,'action.save',By.NAME) # Clica para salvar.
                            acessar_iframe_default(navegador) # Acessa Iframe primario
                    enviarkey_elemento(navegador,'id_txt_dadosDaCobranca__dadosDoFaturamentoVariavel__numeroContratoProtheus__',By.ID,int(planilha.iloc[linha]['NUMERO DO CONTRATO'])) # Envia o numero de contrato
                    clicar_elemento(navegador,'//*[@id="ui-id-10"]/li',By.XPATH) # Clica no numero de contrato
                    clicar_elemento(navegador,'//*[@id="dibButtons"]/input[1]',By.XPATH) # Clica no numero de contrato
                    navegador.switch_to.default_content()#Volta para o inicio

                    # ---------------------- Esta Parte se refere aos Anexos ------------------------
                    enviar_anexo(navegador,linha,'//*[@id="menu_bar_genericoHistoricoAtendimento"]/li[1]','var_dadosDaCobranca__historico__anexo__','//*[@id="progress-complete-var_dadosDaCobranca__historico__anexo__"]/span','var_dadosDaCobranca__historico__registro__')#Envia Anexos
                    enviar_emails(navegador,linha,"//li[@onclick=\"activeDeactiveObjMenu2(this);javascript: ellist_emailClienteFP__.addNewItem('CreateItens', true);\"]//a[@id='createitens']",'var_emailClienteFP__Email__')
                    input('Confirma o lançamento!!!')
                    navegador.switch_to.default_content()
                    clicar_elemento(navegador,'action.send',By.NAME)
                    while len(navegador.find_elements(By.CLASS_NAME, 'alert')) == 0: # Loop para aguardar o alerta carregar!
                        time.sleep(1)
                    time.sleep(1)
                    copiar_linha_ativa(planilha, local_destino, 'Novo', linha,nome_cob)
                    navegador.close() # Fecha a aba apos Alerta Carregar!!
                    navegador.switch_to.window(aba_orignal)
                    time.sleep(1)
                    acessar_iframe_default(navegador)

                elif planilha.iloc[linha]['TIPO'].lower() == "fixo" and planilha.iloc[linha]['RATEIO'].lower() == "não":
                    enviarkey_elemento(navegador,'id_tipoDeMedicao__',By.ID,'Fixa')# Tipo de medição
                    clicar_elemento(navegador,'//*[@id="tab_bar_"]/li[2]/a',By.XPATH) # Clica na guia Dados de cobrança
                    nome_cob = texto_elemento(navegador,'headerTitle',By.ID)
                    enviarkey_elemento(navegador,'id_txt_dadosCobranca__DadosDoCliente__',By.ID,tratar_cnpj(planilha.iloc[linha]['CNPJ']))# Envia CNPJ
                    clicar_elemento(navegador,'ui-menu-item',By.CLASS_NAME) # Clica no CNPJ informado
                    enviarkey_elemento(navegador,'var_dadosCobranca__descricaoServico__',By.ID,planilha.iloc[linha]['DESCRICAO'])# Envia Descrição
                    enviarkey_elemento(navegador,'var_dadosCobranca__dadosParaHistorico2__HouvePrestacaoDeServicos__',By.ID,'Sim')# Prestação de Serviço - Padrão
                    enviarkey_elemento(navegador,'id_txt_dadosCobranca__dadosParaHistorico2__numeroContratoProtheus__',By.ID,str(int(planilha.iloc[linha]['NUMERO DO CONTRATO'])))# Numero de contrato
                    clicar_elemento(navegador,'//*[@id="ui-id-6"]/li',By.XPATH) # Clica no contrato informado
                    enviarkey_elemento(navegador,'id_dadosCobranca__dadosParaHistorico2__diaLimiteNFCliente__',By.ID,str(int(planilha.iloc[linha]['DIA LIMITE'])))# Data Limite
                    opcoes_pagamento(navegador,'//*[@id="mul_dadosCobranca__formaDeEntradaDosRecursosRevisado_ori"]/option[1]','move_this_right_mul_dadosCobranca__formaDeEntradaDosRecursosRevisado') # Loop para opções de pagamento
                    enviarkey_elemento(navegador,'//*[@id="var_dadosCobranca__rateio__"]',By.XPATH,'Não')# Tipo de Rateio

                    clicar_elemento(navegador,'id_dadosCobranca__filaisSemRateio__UOCRProtheus___anchor',By.ID) # Pequisa
                    acessar_iframe_default(navegador)# Acessa o Iframe

                    clicar_elemento(navegador,'vfilter',By.ID) # Clica no Filtro
                    acessar_iframe_default(navegador) # Acessa Iframe do Filtro
                    enviarkey_elemento(navegador,'var_codclvlr__',By.NAME,int(planilha.iloc[linha]['CLASSE DE VALOR'])) # Envia Classa de valor Cliente
                    enviarkey_elemento(navegador,'var_codfilialprotheus__',By.NAME,cod_filial) # Envia COD FILIAL - PADRÃO
                    enviarkey_elemento(navegador,'var_coduo__',By.NAME,cod_uo) # Envia COD UO - PADRÃO
                    enviarkey_elemento(navegador,'var_codccusto__',By.NAME,str(int(planilha.iloc[linha]['CR1']))) # Envia COD PRODUTO
                    clicar_elemento(navegador,'searchbutton',By.ID) # Clica na Pesquisa
                    acessar_iframe_default(navegador) # Acessa Iframe da Pesquisa
                    clicar_elemento(navegador,'tooltip0',By.ID)
                    navegador.switch_to.default_content()

                    valor = float(planilha.iloc[linha]['VALOR1'])
                    parcelas = int(planilha.iloc[linha]['PARCELA'])
                    valor_parcela = valor / parcelas
                    date = planilha.iloc[linha]['DATA']

                    clicar_elemento(navegador,'//*[@id="menu_bar_FinCobFFDatasVencimentos"]/li[2]',By.XPATH) # Itens novos

                    for i in range(int(planilha.iloc[linha]['PARCELA'])):
                            new_data = date + relativedelta(months=i)
                            new_date_str = new_data.strftime('%d/%m/%Y')
                            acessar_iframe_default(navegador) # Acessa Iframe dos itens novos
                            enviarkey_elemento(navegador,'var_dadosCobranca__cobrancas__data__',By.ID,new_date_str) # Envia Data
                            enviarkey_java(navegador,'var_dadosCobranca__cobrancas__valor__',valor_parcela) # Envia Valor
                            clicar_elemento(navegador,'//*[@id="dibButtons"]/input[1]',By.XPATH) # Botão Ok
                    acessar_iframe_default(navegador) # Acessa Iframe dos itens novos
                    clicar_elemento(navegador,'cancelButtonModal',By.ID) # Botão Cancelar
                    navegador.switch_to.default_content()
                    enviarkey_elemento(navegador,'var_dadosCobranca__Observacao__',planilha.iloc[linha]['OBSERVACAO']) # Envia Observação
                    enviar_emails(navegador,linha,"//li[@onclick=\"activeDeactiveObjMenu2(this);javascript: ellist_EmailDeContatoDosClientes__.addNewItem('CreateItens', true);\"]/a[@id='createitens']",'var_EmailDeContatoDosClientes__Email__')
                    input('Confirma o lançamento!!!')
                    navegador.switch_to.default_content()
                    clicar_elemento(navegador,'action.send',By.NAME)
                    while len(navegador.find_elements(By.CLASS_NAME, 'alert')) == 0: # Loop para aguardar o alerta carregar!
                        time.sleep(1)
                    time.sleep(1)
                    copiar_linha_ativa(planilha, local_destino, 'Novo', linha,nome_cob)
                    navegador.close() # Fecha a aba apos Alerta Carregar!!
                    navegador.switch_to.window(aba_orignal)
                    time.sleep(1)
                    acessar_iframe_default(navegador)

                elif planilha.iloc[linha]['TIPO'].lower() == "fixo" and planilha.iloc[linha]['RATEIO'].lower() == "sim":
                    enviarkey_elemento(navegador,'id_tipoDeMedicao__',By.ID,'Fixa')# Tipo de medição
                    clicar_elemento(navegador,'//*[@id="tab_bar_"]/li[2]/a',By.XPATH) # Clica na guia Dados de cobrança
                    nome_cob = texto_elemento(navegador,'headerTitle',By.ID)
                    enviarkey_elemento(navegador,'id_txt_dadosCobranca__DadosDoCliente__',By.ID,tratar_cnpj(planilha.iloc[linha]['CNPJ']))# Envia CNPJ
                    clicar_elemento(navegador,'ui-menu-item',By.CLASS_NAME) # Clica no CNPJ informado
                    enviarkey_elemento(navegador,'var_dadosCobranca__descricaoServico__',By.ID,planilha.iloc[linha]['DESCRICAO'])# Envia Descrição
                    enviarkey_elemento(navegador,'var_dadosCobranca__dadosParaHistorico2__HouvePrestacaoDeServicos__',By.ID,'Sim')# Prestação de Serviço - Padrão
                    enviarkey_elemento(navegador,'id_txt_dadosCobranca__dadosParaHistorico2__numeroContratoProtheus__',By.ID,str(int(planilha.iloc[linha]['NUMERO DO CONTRATO'])))# Numero de contrato
                    clicar_elemento(navegador,'//*[@id="ui-id-6"]/li',By.XPATH) # Clica no contrato informado
                    enviarkey_elemento(navegador,'id_dadosCobranca__dadosParaHistorico2__diaLimiteNFCliente__',By.ID,str(int(planilha.iloc[linha]['DIA LIMITE'])))# Data Limite
                    opcoes_pagamento(navegador,'//*[@id="mul_dadosCobranca__formaDeEntradaDosRecursosRevisado_ori"]/option[1]','move_this_right_mul_dadosCobranca__formaDeEntradaDosRecursosRevisado')# Loop para opções de pagamento
                    enviarkey_elemento(navegador,'//*[@id="var_dadosCobranca__rateio__"]',By.XPATH,'Sim')# Tipo de Rateio
                    parcemlamento = 0
                    for i in range(6):
                        if not pd.isna(planilha.iloc[linha][f'CR{i+1}']):
                            parcemlamento += 1
                    date = planilha.iloc[0]['DATA']
                    for i in range(int(planilha.iloc[linha]['PARCELA'])):
                        clicar_elemento(navegador,'//*[@id="menu_bar_finCobDataXFilialXCcusto"]/li[1]',By.XPATH) # Clica produtos novo
                        acessar_iframe_default(navegador)# Acessa o Iframe
                        new_data = date + relativedelta(months=i)
                        new_date_str = new_data.strftime('%d/%m/%Y')
                        enviarkey_elemento(navegador,'var_dadosCobranca__cobRateio__dataCobranca__',By.ID,new_date_str) # Envia Data
                        contador = 0
                        for i in range(parcemlamento):
                            clicar_elemento(navegador,'//*[@id="menu_bar_finCobFFCentroValor"]/li[1]',By.XPATH) # Pequisa
                            acessar_iframe_default(navegador)# Acessa o Iframe
                            enviarkey_elemento(navegador,'id_txt_dadosCobranca__cobRateio__filialCustos__FilialProtheus__',By.ID,cod_filial) # Envia COD FILIAL - PADRÃO
                            clicar_elemento(navegador,'ui-menu-item',By.CLASS_NAME)
                            enviarkey_elemento(navegador,'id_txt_dadosCobranca__cobRateio__filialCustos__UO__',By.ID,cod_uo) # Envia COD FILIAL - PADRÃO
                            clicar_elemento(navegador,'//*[@id="ui-id-4"]/li',By.XPATH)
                            clicar_elemento(navegador,'id_dadosCobranca__cobRateio__filialCustos__UOXCRProtheus___anchor',By.ID)

                            acessar_iframe_default(navegador)# Acessa o Iframe
                            clicar_elemento(navegador,'vfilter',By.ID) # Clica no Filtro
                            acessar_iframe_default(navegador) # Acessa Iframe do Filtro
                            enviarkey_elemento(navegador,'var_codclvlr__',By.NAME,str(planilha.iloc[linha]['CLASSE DE VALOR'])) # Envia Classa de valor Cliente
                            enviarkey_elemento(navegador,'var_codfilialprotheus__',By.NAME,cod_filial) # Envia COD FILIAL - PADRÃO
                            enviarkey_elemento(navegador,'var_coduo__',By.NAME,cod_uo) # Envia COD UO - PADRÃO
                            enviarkey_elemento(navegador,'var_codccusto__',By.NAME,str(int(planilha.iloc[linha][f'CR{i+1}']))) # Envia COD PRODUTO
                            clicar_elemento(navegador,'searchbutton',By.ID) # Clica na Pesquisa
                            acessar_iframe_default(navegador) # Acessa Iframe da Pesquisa
                            clicar_elemento(navegador,'tooltip0',By.ID)
                            navegador.switch_to.default_content()
                            acessar_iframe_default(navegador) # Acessa Iframe da Pesquisa
                            enviarkey_java(navegador,'var_dadosCobranca__cobRateio__filialCustos__valor__',planilha.iloc[linha][f'VALOR{i+1}']/planilha.iloc[linha]['PARCELA']) # Envia Valor
                            clicar_elemento(navegador,'//*[@id="dibButtons"]/input[1]',By.XPATH) #botão ok
                            acessar_iframe_default(navegador) # Acessa Iframe da Pesquisa
                            contador += 1
                        for i in range(contador):
                            clicar_elemento_rustico(navegador,f'//*[@id="{i}"]/td[2]',By.XPATH) #Para acertar as porcentagens 
                            acessar_iframe_default(navegador) # Acessa Iframe da Pesquisa
                            clicar_elemento_rustico(navegador,'//*[@id="dibButtons"]/input[1]',By.XPATH) #botão ok
                            acessar_iframe_default(navegador) # Acessa Iframe da Pesquisa
                        clicar_elemento(navegador,'//*[@id="dibButtons"]/input[1]',By.XPATH) #botão ok
                        navegador.switch_to.default_content()

                    enviarkey_elemento(navegador,'var_dadosCobranca__Observacao__',planilha.iloc[linha]['OBSERVACAO']) # Envia Observação
                    enviar_emails(navegador,linha,"//li[@onclick=\"activeDeactiveObjMenu2(this);javascript: ellist_EmailDeContatoDosClientes__.addNewItem('CreateItens', true);\"]/a[@id='createitens']",'var_EmailDeContatoDosClientes__Email__')
                    input('Confirma o lançamento!!!')
                    navegador.switch_to.default_content()
                    clicar_elemento(navegador,'action.send',By.NAME)
                    while len(navegador.find_elements(By.CLASS_NAME, 'alert')) == 0: # Loop para aguardar o alerta carregar!
                        time.sleep(1)
                    time.sleep(1)
                    copiar_linha_ativa(planilha, local_destino, 'Novo', linha,nome_cob)
                    navegador.close() # Fecha a aba apos Alerta Carregar!!
                    navegador.switch_to.window(aba_orignal)
                    time.sleep(1)
                    acessar_iframe_default(navegador)

            copiar_para_planilha(planilha_destino,local_destino)
        
        elif escolha == 0:
            print("Saindo...")
            break
        
        else:
            print("Escolha inválida. Tente novamente.")
    except ValueError:
        print("Entrada inválida. Digite um número.")
    except Exception as e:
        print(f"Ocorreu um erro: {e}")
        copiar_para_planilha(planilha_destino,local_destino)
        input('Chame a T.I')