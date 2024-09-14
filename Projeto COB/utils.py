from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from tkinter.filedialog import askopenfilename
from datetime import datetime
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from tkinter import messagebox
import tkinter as tk
import calendar
import time
import locale
import pandas as pd
import shutil
import os

locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')

def acessar_iframe(nav, tempo_espera):
    time.sleep(tempo_espera)# Espera antes de mudar para o Iframe
    iframe = WebDriverWait(nav, 180).until(EC.presence_of_element_located((By.TAG_NAME, 'iframe'))) # Espera 180 segundos até o iframe aparecer!
    nav.switch_to.frame(iframe) #Troca para o iframe

def acessar_iframe_default(nav, tempo_espera, timeout=10):
    time.sleep(tempo_espera)# Espera antes de mudar para o conteúdo padrão
    nav.switch_to.default_content()
    iframe = WebDriverWait(nav, timeout).until(EC.presence_of_element_located((By.TAG_NAME, 'iframe')))# Espera até que o iframe esteja presente
    nav.switch_to.frame(iframe)# Troca para o iframe 

def clicar_elemento(nav,elemento,tipo):
    try:
        obj = WebDriverWait(nav, 30).until(EC.presence_of_element_located((tipo, elemento))) # Aguarda 180 segundos até o elemento carregar
        nav.execute_script("arguments[0].click();", obj) # Clica no objeto utilizando paramentros JavaScript.
    except Exception:
        root = tk.Tk()
        root.withdraw()  # Oculta a janela principal do Tkinter
        messagebox.showwarning("Alerta", "Botão ou campo não encontrado. Localize e preencha manualmente. O código continuará a executar ao apertar OK.")
        root.destroy()

def clicar_elemento_rustico(nav,elemento,tipo):
    try:
        WebDriverWait(nav, 60).until(EC.presence_of_element_located((tipo, elemento))) # Aguarda 180 segundos até o elemento carregar
        nav.find_element(tipo, elemento).click()# Clicar na Pesquisa da forma tradicional do Selenium!
    except Exception:
        root = tk.Tk()
        root.withdraw()  # Oculta a janela principal do Tkinter
        messagebox.showwarning("Alerta", "Botão ou campo não encontrado. Localize e preencha manualmente. O código continuará a executar ao apertar OK.")
        root.destroy()

def enviarkey_elemento(nav,elemento,tipo,texto):
    try:
        WebDriverWait(nav, 60).until(EC.presence_of_element_located((tipo, elemento))) # Aguarda 180 segundos até o elemento carregar
        nav.find_element(tipo, elemento).send_keys(texto) # envia dados para o elemento
    except Exception:
        root = tk.Tk()
        root.withdraw()  # Oculta a janela principal do Tkinter
        messagebox.showwarning("Alerta", "Botão ou campo não encontrado. Localize e preencha manualmente. O código continuará a executar ao apertar OK.")
        root.destroy()

def primeiro_e_ultimo_dia_do_mes(ano, mes):
    primeiro_dia = datetime(ano, mes, 1)# Primeiro dia do mês
    ultimo_dia = datetime(ano, mes, calendar.monthrange(ano, mes)[1])# Último dia do mês
    # Formatar as datas no formato dd/mm/aaaa
    primeiro_dia_formatado = primeiro_dia.strftime('%d/%m/%Y')
    ultimo_dia_formatado = ultimo_dia.strftime('%d/%m/%Y')
    return primeiro_dia_formatado, ultimo_dia_formatado

def copiar_linha_ativa(df, destino, sheet_name, linha, texto_adicional=None):
    linha_ativa = df.iloc[[linha]].dropna(how='all')  # Seleciona a linha ativa específica (não vazia)
    book = load_workbook(destino)  # Tenta carregar a planilha de destino existente:
    sheet = book[sheet_name]
    next_row = sheet.max_row + 1# Encontra a próxima linha vazia na planilha de destino
    for r_idx, row in enumerate(dataframe_to_rows(linha_ativa, index=False, header=False), start=next_row):# Adiciona a linha ativa à planilha de destino
        for c_idx, value in enumerate(row, 1):
            sheet.cell(row=r_idx, column=c_idx, value=value)
    if sheet_name == 'Novo':
        cob_column_index = 29  # Index da coluna COB, alterar manualmente caso planilha seja modificada!!
        sheet.cell(row=next_row, column=cob_column_index, value=texto_adicional)# Adiciona o texto na última coluna da nova linha
        hoje = datetime.today().strftime('%d/%m/%Y')# Pega a data de hoje
        sheet.cell(row=next_row, column=cob_column_index + 1, value=hoje)# Adiciona a data de hoje na coluna seguinte
    else:
        dia_column_index = 25
        hoje = datetime.today().strftime('%d/%m/%Y')# Pega a data de hoje
        sheet.cell(row=next_row, column=dia_column_index, value=hoje)# Adiciona a data de hoje na coluna  
    
    book.save(destino)# Salva o arquivo de destino

def copiar_para_planilha(local_destino, local_origem):
        os.makedirs(os.path.dirname(local_destino), exist_ok=True)
        shutil.copy2(local_origem, local_destino)

def enviarkey_java(nav, element_name, value):
    try:
        WebDriverWait(nav, 60).until(EC.presence_of_element_located((By.NAME, element_name)))
        script = f"document.getElementsByName('{element_name}')[0].value='{value}';" # Simula a entrada de dados via JavaScript para evitar interferência da máscara de entrada
        nav.execute_script(script)
        script = f"""
        var input = document.getElementsByName('{element_name}')[0];
        var event = new Event('input', {{ bubbles: true }});
        input.dispatchEvent(event);
        """
        nav.execute_script(script) # Dispara eventos para que o script de máscara possa processar o novo valor
    except Exception:
        root = tk.Tk()
        root.withdraw()  # Oculta a janela principal do Tkinter
        messagebox.showwarning("Alerta", "Botão ou campo não encontrado. Localize e preencha manualmente. O código continuará a executar ao apertar OK.")
        root.destroy()

def selecionar_arquivo():
    caminho_arquivo = askopenfilename(title="Selecione a Planilha COB!") # Solciita o usuario selecionar a planilha!
    return caminho_arquivo

def texto_elemento(nav,elemento,tipo):
    obj = WebDriverWait(nav, 60).until(EC.presence_of_element_located((tipo, elemento))) # Aguarda 60 segundos até o elemento carregar
    texto = obj.text
    return texto.split('-')[0]

def esperar_elementos_carregar(nav, timeout=60):
        WebDriverWait(nav, timeout).until(
            lambda driver: len(driver.find_elements(By.CLASS_NAME, "item")) > 0 or 
                           len(driver.find_elements(By.XPATH, '//div[contains(@class, "no-results-default-boxes") and contains(@class, "ng-scope") and contains(., "Sua Caixa de Entrada está vazia")]')) > 0
        )

def enviar_emails(nav, linha,click,campo,planilha):
    email = planilha.iloc[linha]['EMAILS'] # recebe e-mails da planilha (Os mesmo devem ser separados por uma '/')
    lst_email = email.split('/') # Transforma os e-mails recebidos em lista, separador '/'
    clicar_elemento(nav,click,By.XPATH) # Itens novos e-mails
    for i in range(len(lst_email)):
            acessar_iframe_default(nav) # Acessa Iframe dos e-mails
            enviarkey_elemento(nav,campo,By.ID,lst_email[i]) # Envia e-mail
            clicar_elemento_rustico(nav,'form_container',By.ID) # Clica no container para o e-mail carregar
            clicar_elemento(nav,'//*[@id="dibButtons"]/input[1]',By.XPATH) # Botão Ok
    acessar_iframe_default(nav) # Acessa Iframe dos e-mails
    clicar_elemento(nav,'cancelButtonModal',By.ID) # Botão Cancelar, para fechar janelas!

def tratar_cnpj(cnpj):
    cnpj = str(cnpj)  # Converte o CNPJ para string, caso não seja
    if '.' in cnpj or '-' in cnpj or '/' in cnpj:# Verifica se o CNPJ contém '.', '-', '/'
        cnpj_limpo = cnpj.replace('.', '').replace('/', '').replace('-', '')# Remove pontuações se existirem
    else:
        cnpj_limpo = cnpj  # Se não tiver pontuações, já está limpo
    cnpj_formatado = str(int(cnpj_limpo))# Converte para inteiro para remover zeros à esquerda, depois para string novamente
    return cnpj_formatado

def enviar_anexo(nav,linha,click,element,status,descr,planilha):
    for anexo in range(2): 
        if pd.isna(planilha.iloc[linha][f'ARQUIVO{anexo+1}']):
            pass
        else:
            clicar_elemento(nav,click,By.XPATH) # Clica no anexo para enviar arquivo
            acessar_iframe(nav)#Acesso Iframe
            enviarkey_elemento(nav,element,By.ID,fr"{planilha[:-16]}Arquivos\{planilha.iloc[linha][f'ARQUIVO{anexo+1}']}") # Envia o anexo
            while len(nav.find_elements(By.XPATH, status)) == 0: # Loop para aguardar a lista de itens carregar, se a lista não carregar a pesquisa não funciona!
                time.sleep(1)
            time.sleep(1)
            enviarkey_elemento(nav,descr,By.ID,planilha.iloc[linha][f'DESCRICAO{anexo+1}']) # Envia a descrição
            clicar_elemento(nav,'//*[@id="dibButtons"]/input[1]',By.XPATH) 
            nav.switch_to.default_content()#Volta para o inicio

def opcoes_pagamento(nav,selec,seta):
    for i in range(2):#Loop para selecionar as opções de pagamento
        nav.find_element(By.XPATH, selec).click()
        nav.find_element(By.ID, seta).click()  

def clicar_porcentagem(nav,contador,linha,planilha):
    for i in range(contador): # Baseado na soma do Contador clica nos itens
        clicar_elemento_rustico(nav,f'//*[@id="{i}"]/td[2]',By.XPATH) # Clica no Item baseado nos indices (No fusion o indice 0 conta!)!!
        acessar_iframe_default(nav) # Acessa Iframe primario
        clicar_elemento(nav,'action.save',By.NAME) # Clica para salvar.
        acessar_iframe_default(nav) # Acessa Iframe primario
    enviarkey_elemento(nav,'id_txt_dadosDaCobranca__dadosDoFaturamentoVariavel__numeroContratoProtheus__',By.ID,str(int(planilha.iloc[linha]['NUMERO DO CONTRATO']))) # Envia o numero de contrato
    clicar_elemento(nav,'//*[@id="ui-id-10"]/li',By.XPATH) # Clica no numero de contrato
    clicar_elemento(nav,'action.save',By.NAME) # Clica para salvar.
    nav.switch_to.default_content()#Volta para o inicio

def dados_rateio(nav,linha,cod_filial,cod_uo,planilha):
    clicar_elemento(nav,'//*[@id="menu_bar_FINFFCobFaturamentoVariavelCentroDeResultadosXFilialXValor"]/li[1]',By.XPATH) # Clica para abrir campo de produtos
    acessar_iframe_default(nav) # Acessa Iframe da Pesquisa de produtos
    opcoes_pagamento(nav,'//*[@id="mul_dadosDaCobranca__dadosDoFaturamentoVariavel__dadosDoRateio__formaDeEntradaDosRecursos_ori"]/option[1]','move_this_right_mul_dadosDaCobranca__dadosDoFaturamentoVariavel__dadosDoRateio__formaDeEntradaDosRecursos')#Loop para selecionar as opções de pagamento
    clicar_elemento(nav,'id_dadosDaCobranca__dadosDoFaturamentoVariavel__dadosDoRateio__UOCRProtheus___anchor',By.ID) # Clica para abrir campo de pesquisa
    acessar_iframe_default(nav) # Acessa Iframe da Pesquisa
    clicar_elemento(nav,'//*[@id="menu_bar_EXTERNOProtheusAmarracaoContabil"]/li',By.XPATH) # Clica para abrir filtro
    acessar_iframe_default(nav) # Acessa Iframe do Filtro
    enviarkey_elemento(nav,'var_codclvlr__',By.NAME,str(int(planilha.iloc[linha]['CLASSE DE VALOR']))) # Envia Classa de valor Cliente
    enviarkey_elemento(nav,'var_codfilialprotheus__',By.NAME,cod_filial) # Envia COD FILIAL - PADRÃO
    enviarkey_elemento(nav,'var_coduo__',By.NAME,cod_uo) # Envia COD UO - PADRÃO

def iniciar_navegador():
    options = webdriver.ChromeOptions()
    options.add_experimental_option("detach", True)# Para o mesmo não fechar apos execução
    navegador = webdriver.Chrome(options=options)# Executa o navegador
    navegador.get('https://fusion.fiemg.com.br/fusion/portal')
    navegador.maximize_window()# Maximiza a janela do navegador
    return navegador
    
def variavel_novo(nav,linha,planilha,primeiro_dia,ultimo_dia):
        enviarkey_elemento(nav,'id_tipoDeMedicao__',By.ID,'Variavel')# Tipo de medição
        enviarkey_elemento(nav,'var_dadosDaCobranca__APeriodicidadeDoFaturamentoEMensal__',By.ID,'Sim')# Periodicidade
        enviarkey_elemento(nav,'var_dadosDaCobranca__dadosParaHistorico__HouvePrestacaoDeServicos__',By.ID,'Sim')# Prestação de serviço
        enviarkey_elemento(nav,'id_txt_dadosDaCobranca__dadosParaHistorico__numeroContratoProtheus__',By.ID,str(int(planilha.iloc[linha]['NUMERO DO CONTRATO'])))# Numero do Contrato
        clicar_elemento(nav,'//*[@id="ui-id-3"]/li',By.XPATH) # Clica no Numero do contrato
        enviarkey_elemento(nav,'var_dadosDaCobranca__dadosParaHistorico__existeGasOuOS__',By.ID,'Não')# Existe GAS ou OS
        enviarkey_elemento(nav,'var_dadosDaCobranca__dadosParaHistorico__parcelaContrato__',By.ID,'0')# Informa Parcela
        enviarkey_elemento(nav,'var_dadosDaCobranca__dadosParaHistorico__inicioPrestacao__',By.ID,primeiro_dia)# Data Inicio
        enviarkey_elemento(nav,'var_dadosDaCobranca__dadosParaHistorico__finPrestacao__',By.ID,ultimo_dia)# Data Fim
        enviarkey_elemento(nav,'id_dadosDaCobranca__dadosParaHistorico__diaLimiteNFCliente__',By.ID,str(int(planilha.iloc[linha]['DIA LIMITE'])))# Dia Limite
        enviarkey_elemento(nav,'var_dadosDaCobranca__cobrancaRelacionadaComConvenio__',By.ID,'Não')# Convenio

def esperar_alerta(nav, cob, aba_original,planilha, local_destino,nome_guia,linha,texto_adicional=None):
    alert = WebDriverWait(nav, 30).until(EC.presence_of_element_located((By.CSS_SELECTOR, '.alert')))# Espera que qualquer alerta de sucesso ou perigo apareça
    if 'alert-success' in alert.get_attribute('class'):# Verifica a classe do alerta
        if nome_guia == 'Medição':
            copiar_linha_ativa(planilha, local_destino, 'Medição', linha)
        elif nome_guia == 'Novo':
            copiar_linha_ativa(planilha, local_destino, 'Novo', linha,texto_adicional)
        print('Cob executado com sucesso!!')
        nav.close()  # Fecha a aba após o alerta carregar
        nav.switch_to.window(aba_original)  # Volta para a aba original
    else:
        print(f"O {cob} apresentou falha ao enviar, não foi incluído na planilha Histórico!!")
        nav.close()  # Fecha a aba após o alerta carregar
        nav.switch_to.window(aba_original)  # Volta para a aba original

